"""Сервисы отчетности и экспорта для provider location admin."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any, DefaultDict, cast

import openpyxl
from django.db.models import Count, Max, Min, Q
from django.http import HttpResponse
from django.utils import timezone, translation
from django.utils.dateparse import parse_date
from django.utils.formats import date_format
from django.utils.translation import gettext_lazy as _
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from booking.constants import (
    BOOKING_STATUS_ACTIVE,
    BOOKING_STATUS_CANCELLED,
    BOOKING_STATUS_COMPLETED,
    COMPLETED_BY_SYSTEM,
)
from booking.manual_v2_models import ManualBooking
from booking.models import Booking
from billing.models import Invoice, InvoiceLine, PaymentHistory
from catalog.models import Service
from pets.models import PetType
from providers.models import (
    Employee,
    EmployeeLocationRole,
    EmployeeLocationService,
    EmployeeProvider,
    ProviderLocation,
    ProviderLocationService,
    Schedule,
)
from scheduling.models import SickLeave, Vacation


class ProviderLocationReportError(Exception):
    """Ошибка валидации параметров отчетности."""


@dataclass(frozen=True)
class StaffRosterEntry:
    """Нормализованная строка штатного состава филиала."""

    employee_id: int
    location_id: int
    location_name: str
    full_name: str
    email: str
    is_active: bool
    hire_date: date | None
    dismissal_datetime: datetime | None
    service_names: tuple[str, ...]
    service_summary: str
    service_count: int


class ProviderLocationReportingService:
    """Собирает отчеты филиала/организации и формирует XLSX-выгрузки."""

    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_COLUMNS = 16_384
    PERIOD_REPORT_CODES = frozenset({
        'staff_schedule',
        'staff_load',
        'staff_performance',
        'bookings_summary',
        'financial_revenue',
        'platform_settlement',
    })
    REPORT_CODES = frozenset({
        'staff_roster',
        'staff_schedule',
        'services_price_list',
        'service_coverage',
        'staff_load',
        'staff_performance',
        'bookings_summary',
        'financial_revenue',
        'platform_settlement',
    })
    FIELD_LABELS = {
        'metric': _('Metric'),
        'value': _('Value'),
        'total_services': _('Total services'),
        'covered_services': _('Covered services'),
        'uncovered_services': _('Uncovered services'),
        'category_name': _('Category'),
        'services': _('Services'),
        'covered': _('Covered'),
        'uncovered': _('Uncovered'),
        'location_name': _('Branch'),
        'service_name': _('Service'),
        'service_path': _('Service path'),
        'top_category_name': _('Top category'),
        'covered_staff_count': _('Assigned staff'),
        'is_covered': _('Covered'),
        'active_staff': _('Active staff'),
        'available_minutes': _('Available minutes'),
        'busy_minutes': _('Busy minutes'),
        'free_minutes': _('Free minutes'),
        'utilization_percent': _('Utilization %'),
        'completed_bookings': _('Completed bookings'),
        'revenue': _('Revenue'),
        'employee_name': _('Employee'),
        'employees': _('Employees'),
        'completed_services': _('Completed services'),
        'completed_count': _('Completed count'),
        'total_bookings': _('Total bookings'),
        'total_amount': _('Total amount'),
        'auto_completed_count': _('Auto-completed'),
        'overdue_not_completed_count': _('Overdue not completed'),
        'source': _('Source'),
        'status': _('Status'),
        'count': _('Count'),
        'amount': _('Amount'),
        'start_time': _('Start time'),
        'completed_by_actor': _('Completed by'),
        'is_auto_completed': _('Auto-completed'),
        'total_revenue': _('Total revenue'),
        'average_ticket': _('Average ticket'),
        'invoice_number': _('Invoice number'),
        'invoice_status': _('Invoice status'),
        'issued_at': _('Issued at'),
        'due_date': _('Due date'),
        'payment_date': _('Payment date'),
        'invoiced_amount': _('Invoiced amount'),
        'invoiced_commission': _('Invoiced commission'),
        'invoiced_total_with_vat': _('Invoiced total with VAT'),
        'paid_amount': _('Paid amount'),
        'outstanding_amount': _('Outstanding amount'),
        'overdue_amount': _('Overdue amount'),
        'commission': _('Commission'),
        'rate': _('Rate'),
        'vat_amount': _('VAT amount'),
        'total_with_vat': _('Total with VAT'),
        'booking_code': _('Booking code'),
        'invoice_count': _('Invoice count'),
        'payment_count': _('Payment count'),
        'full_name': _('Full name'),
        'email': _('Email'),
        'is_active': _('Active'),
        'hire_date': _('Hire date'),
        'dismissal_date': _('Dismissal date'),
        'dismissed_at': _('Dismissal date'),
        'service_summary': _('Services'),
        'service_count': _('Service count'),
        'day_label': _('Day'),
        'total_staff': _('Total staff'),
        'dismissed_staff': _('Dismissed staff'),
    }
    SHEET_TITLES = {
        'Summary': _('Summary'),
        'Staff': _('Staff'),
        'Services': _('Services'),
        'Coverage': _('Coverage'),
        'Uncovered': _('Uncovered'),
        'Categories': _('Categories'),
        'Schedule': _('Schedule'),
        'Employees': _('Employees'),
        'Locations': _('Locations'),
        'Details': _('Details'),
        'Sources': _('Sources'),
        'Statuses': _('Statuses'),
        'Bookings': _('Bookings'),
        'Data': _('Data'),
        'Invoices': _('Invoices'),
        'Payments': _('Payments'),
    }
    RUSSIAN_FIELD_LABELS = {
        'metric': 'Метрика',
        'value': 'Значение',
        'total_services': 'Всего услуг',
        'covered_services': 'Покрыто услуг',
        'uncovered_services': 'Непокрыто услуг',
        'category_name': 'Категория',
        'services': 'Услуги',
        'covered': 'Покрыто',
        'uncovered': 'Не покрыто',
        'location_name': 'Филиал',
        'service_name': 'Услуга',
        'service_path': 'Путь услуги',
        'top_category_name': 'Верхняя категория',
        'covered_staff_count': 'Назначено сотрудников',
        'is_covered': 'Покрыто',
        'active_staff': 'Активный персонал',
        'available_minutes': 'Доступно минут',
        'busy_minutes': 'Занято минут',
        'free_minutes': 'Свободно минут',
        'utilization_percent': 'Загрузка, %',
        'completed_bookings': 'Завершено записей',
        'revenue': 'Выручка',
        'employee_name': 'Сотрудник',
        'employees': 'Сотрудники',
        'completed_services': 'Оказано услуг',
        'completed_count': 'Количество выполнений',
        'total_bookings': 'Всего записей',
        'total_amount': 'Общая сумма',
        'auto_completed_count': 'Автозавершено',
        'overdue_not_completed_count': 'Просрочено и не завершено',
        'source': 'Источник',
        'status': 'Статус',
        'count': 'Количество',
        'amount': 'Сумма',
        'start_time': 'Время начала',
        'completed_by_actor': 'Завершено кем',
        'is_auto_completed': 'Автозавершение',
        'total_revenue': 'Общая выручка',
        'average_ticket': 'Средний чек',
        'invoice_number': 'Номер счета',
        'invoice_status': 'Статус счета',
        'issued_at': 'Выставлен',
        'due_date': 'Срок оплаты',
        'payment_date': 'Дата оплаты',
        'invoiced_amount': 'Сумма по счетам',
        'invoiced_commission': 'Комиссия по счетам',
        'invoiced_total_with_vat': 'Комиссия с НДС',
        'paid_amount': 'Оплачено',
        'outstanding_amount': 'Остаток к оплате',
        'overdue_amount': 'Просрочено',
        'commission': 'Комиссия',
        'rate': 'Ставка',
        'vat_amount': 'НДС',
        'total_with_vat': 'Итого с НДС',
        'booking_code': 'Код записи',
        'invoice_count': 'Количество счетов',
        'payment_count': 'Количество оплат',
        'full_name': 'ФИО',
        'email': 'Email',
        'is_active': 'Активен',
        'hire_date': 'Дата приема',
        'dismissal_date': 'Дата увольнения',
        'dismissed_at': 'Дата увольнения',
        'service_summary': 'Услуги',
        'service_count': 'Количество услуг',
        'day_label': 'День',
        'total_staff': 'Всего сотрудников',
        'dismissed_staff': 'Уволено',
    }
    RUSSIAN_SHEET_TITLES = {
        'Summary': 'Краткая статистика',
        'Staff': 'Персонал',
        'Services': 'Услуги',
        'Coverage': 'Покрытие',
        'Uncovered': 'Непокрытые',
        'Categories': 'Категории',
        'Schedule': 'График',
        'Employees': 'Сотрудники',
        'Locations': 'Филиалы',
        'Details': 'Детализация',
        'Sources': 'Источники',
        'Statuses': 'Статусы',
        'Bookings': 'Записи',
        'Data': 'Данные',
        'Invoices': 'Счета',
        'Payments': 'Платежи',
    }

    def __init__(
        self,
        *,
        location: ProviderLocation,
        scope: str,
        language_code: str,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> None:
        """Подготавливает общий контекст отчета."""
        self.location = location
        self.provider = location.provider
        self.scope = scope if scope in {'location', 'provider'} else 'location'
        self.language_code = (language_code or 'en').split('-')[0]
        if self.language_code == 'cnr':
            self.language_code = 'me'

        if self.scope == 'provider':
            self.locations = list(
                self.provider.locations.all().order_by('name').only('id', 'name', 'provider_id')
            )
        else:
            self.locations = [self.location]
        self.location_ids = [item.id for item in self.locations]
        self.location_map = {item.id: item for item in self.locations}
        self.start_date = start_date
        self.end_date = end_date
        self.today = timezone.localdate()
        self._service_map: dict[int, Service] | None = None
        self._pet_type_map: dict[int, PetType] | None = None
        self._staff_roster: list[StaffRosterEntry] | None = None

    def build_report(self, report_code: str) -> dict[str, Any]:
        """Возвращает JSON-представление выбранного отчета."""
        if report_code not in self.REPORT_CODES:
            raise ProviderLocationReportError(_('Unknown report code.'))

        if report_code in self.PERIOD_REPORT_CODES:
            start_date, end_date = self._require_period(report_code)
        else:
            start_date, end_date = self.start_date, self.end_date

        builder = getattr(self, f'_build_{report_code}_report')
        payload = builder(start_date=start_date, end_date=end_date)
        return {
            'report_code': report_code,
            'scope': self.scope,
            'generated_at': timezone.now().isoformat(),
            'permissions': {
                'can_view_reports': True,
                'allowed_report_codes': sorted(self.REPORT_CODES),
            },
            'period': {
                'start_date': start_date.isoformat() if start_date else None,
                'end_date': end_date.isoformat() if end_date else None,
            },
            'data': payload,
        }

    def build_xlsx_response(self, report_code: str) -> HttpResponse:
        """Готовит XLSX-файл для выбранного отчета."""
        report_payload = self.build_report(report_code)
        workbook = self._build_workbook(report_payload)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = self.build_xlsx_filename(report_code)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def build_xlsx_bytes(self, report_code: str) -> bytes:
        """Возвращает бинарное содержимое XLSX без HTTP-обертки."""
        report_payload = self.build_report(report_code)
        workbook = self._build_workbook(report_payload)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def build_xlsx_filename(self, report_code: str) -> str:
        """Строит имя XLSX-файла с учетом scope."""
        filename_scope = 'organization' if self.scope == 'provider' else 'location'
        return f'{report_code}-{filename_scope}-{timezone.now().strftime("%Y%m%d-%H%M%S")}.xlsx'

    def _require_period(self, report_code: str) -> tuple[date, date]:
        """Возвращает диапазон периода или выставляет разумный default."""
        if self.start_date and self.end_date:
            if self.start_date > self.end_date:
                raise ProviderLocationReportError(_('start_date must not be later than end_date.'))
            return self.start_date, self.end_date

        if report_code == 'staff_schedule':
            start_date = self.today
            end_date = self.today + timedelta(days=13)
        else:
            end_date = self.today
            start_date = end_date - timedelta(days=29)
        return start_date, end_date

    def _get_service_map(self) -> dict[int, Service]:
        """Загружает услуги иерархии для локализованного отображения."""
        if self._service_map is None:
            self._service_map = {
                item.id: item
                for item in Service.objects.all().only(
                    'id',
                    'name',
                    'name_en',
                    'name_ru',
                    'name_me',
                    'name_de',
                    'parent_id',
                    'search_keywords',
                )
            }
        return self._service_map

    def _get_pet_type_map(self) -> dict[int, PetType]:
        """Загружает типы животных для выгрузок прайса."""
        if self._pet_type_map is None:
            self._pet_type_map = {
                item.id: item
                for item in PetType.objects.all().only(
                    'id',
                    'name',
                    'name_en',
                    'name_ru',
                    'name_me',
                    'name_de',
                    'code',
                )
            }
        return self._pet_type_map

    def _get_localized_name(self, obj: Any, *, fallback_attr: str = 'name') -> str:
        """Возвращает локализованное имя сущности по текущей локали."""
        lang_attr = f'{fallback_attr}_{self.language_code}'
        value = getattr(obj, lang_attr, None) or getattr(obj, fallback_attr, None)
        if value:
            return str(value)
        code = getattr(obj, 'code', None)
        return str(code or '')

    def _get_service_name(self, service_id: int | None) -> str:
        """Возвращает локализованное название услуги."""
        if service_id is None:
            return ''
        service = self._get_service_map().get(service_id)
        if not service:
            return str(service_id)
        return self._get_localized_name(service)

    def _get_pet_type_name(self, pet_type_id: int | None) -> str:
        """Возвращает локализованное название типа животного."""
        if pet_type_id is None:
            return ''
        pet_type = self._get_pet_type_map().get(pet_type_id)
        if not pet_type:
            return str(pet_type_id)
        return self._get_localized_name(pet_type)

    def _get_service_lineage(self, service_id: int | None) -> list[Service]:
        """Строит путь услуги по иерархии вверх."""
        if service_id is None:
            return []
        service_map = self._get_service_map()
        lineage: list[Service] = []
        current = service_map.get(service_id)
        visited: set[int] = set()
        while current and current.id not in visited:
            visited.add(current.id)
            lineage.append(current)
            current = service_map.get(current.parent_id) if current.parent_id else None
        lineage.reverse()
        return lineage

    def _get_top_category_name(self, service_id: int | None) -> str:
        """Возвращает верхний узел иерархии услуги."""
        lineage = self._get_service_lineage(service_id)
        if not lineage:
            return ''
        return self._get_localized_name(lineage[0])

    def _get_service_path(self, service_id: int | None) -> str:
        """Возвращает человекочитаемый путь услуги."""
        lineage = self._get_service_lineage(service_id)
        return ' / '.join(self._get_localized_name(node) for node in lineage)

    def _get_staff_roster_entries(self) -> list[StaffRosterEntry]:
        """Собирает и кэширует штатный состав в текущем scope."""
        if self._staff_roster is not None:
            return self._staff_roster

        employee_ids = list(
            Employee.objects.filter(
                locations__in=self.locations,
                employeeprovider_set__provider=self.provider,
            )
            .distinct()
            .values_list('id', flat=True)
        )
        if not employee_ids:
            self._staff_roster = []
            return self._staff_roster

        employees = {
            item.id: item
            for item in Employee.objects.filter(id__in=employee_ids).select_related('user')
        }
        employment_stats: dict[int, dict[str, Any]] = {}
        for row in (
            EmployeeProvider.objects.filter(
                provider=self.provider,
                employee_id__in=employee_ids,
            )
            .values('employee_id')
            .annotate(
                hire_date=Min('start_date'),
                last_end_date=Max('end_date'),
                active_links=Count('id', filter=Q(end_date__isnull=True) | Q(end_date__gte=self.today)),
            )
        ):
            employment_stats[row['employee_id']] = row

        role_map = {
            (item.employee_id, item.provider_location_id): item
            for item in EmployeeLocationRole.objects.filter(
                employee_id__in=employee_ids,
                provider_location_id__in=self.location_ids,
            ).only('employee_id', 'provider_location_id', 'is_active', 'end_date')
        }

        location_assignments = list(
            Employee.objects.filter(id__in=employee_ids, locations__in=self.locations)
            .values_list('id', 'locations')
            .distinct()
        )

        location_service_name_map: dict[int, dict[int, str]] = defaultdict(dict)
        for item in ProviderLocationService.objects.filter(
            location_id__in=self.location_ids,
            is_active=True,
        ).select_related('service'):
            location_service_name_map[item.location_id][item.service_id] = self._get_localized_name(item.service)

        employee_service_names: dict[tuple[int, int], list[str]] = defaultdict(list)
        employee_service_counts: dict[tuple[int, int], int] = defaultdict(int)
        for item in (
            EmployeeLocationService.objects.filter(
                employee_id__in=employee_ids,
                provider_location_id__in=self.location_ids,
            )
            .select_related('service')
            .order_by('service__name')
        ):
            key = (item.employee_id, item.provider_location_id)
            employee_service_counts[key] += 1
            employee_service_names[key].append(self._get_localized_name(item.service))

        roster_entries: list[StaffRosterEntry] = []
        for employee_id, location_id in location_assignments:
            employee = employees.get(employee_id)
            if employee is None or employee.user is None:
                continue

            location = self.location_map.get(location_id)
            if location is None:
                continue

            employment = employment_stats.get(employee_id, {})
            role_obj = role_map.get((employee_id, location_id))
            has_active_provider_link = bool(employment.get('active_links'))
            is_active = role_obj.is_active if role_obj is not None else has_active_provider_link
            dismissal_datetime = role_obj.end_date if role_obj and role_obj.end_date else None
            if dismissal_datetime is None and employment.get('last_end_date') and not has_active_provider_link:
                dismissal_datetime = datetime.combine(
                    employment['last_end_date'],
                    time.min,
                    tzinfo=timezone.get_current_timezone(),
                )

            key = (employee_id, location_id)
            location_service_count = len(location_service_name_map.get(location_id, {}))
            service_names = tuple(employee_service_names.get(key, []))
            service_count = employee_service_counts.get(key, 0)
            if service_count == 0:
                service_summary = str(_('No assigned services'))
            elif location_service_count and service_count >= location_service_count:
                service_summary = str(_('All branch services'))
            elif service_count <= 3:
                service_summary = ', '.join(service_names)
            else:
                hidden_count = service_count - 3
                service_summary = _('%(visible)s (+%(hidden)s more)') % {
                    'visible': ', '.join(service_names[:3]),
                    'hidden': hidden_count,
                }

            roster_entries.append(
                StaffRosterEntry(
                    employee_id=employee_id,
                    location_id=location_id,
                    location_name=location.name,
                    full_name=(employee.user.get_full_name() or employee.user.email or str(employee_id)).strip(),
                    email=employee.user.email or '',
                    is_active=is_active,
                    hire_date=employment.get('hire_date'),
                    dismissal_datetime=dismissal_datetime,
                    service_names=service_names,
                    service_summary=str(service_summary),
                    service_count=service_count,
                )
            )

        roster_entries.sort(key=lambda item: (item.location_name.casefold(), item.full_name.casefold()))
        self._staff_roster = roster_entries
        return self._staff_roster

    def _iter_dates(self, start_date: date, end_date: date) -> list[date]:
        """Строит список дат периода."""
        total_days = (end_date - start_date).days + 1
        return [start_date + timedelta(days=index) for index in range(max(total_days, 0))]

    def _duration_minutes(
        self,
        start_time: time | None,
        end_time: time | None,
        break_start: time | None = None,
        break_end: time | None = None,
    ) -> int:
        """Считает длительность смены в минутах."""
        if not start_time or not end_time:
            return 0
        start_dt = datetime.combine(self.today, start_time)
        end_dt = datetime.combine(self.today, end_time)
        minutes = max(int((end_dt - start_dt).total_seconds() // 60), 0)
        if break_start and break_end:
            break_start_dt = datetime.combine(self.today, break_start)
            break_end_dt = datetime.combine(self.today, break_end)
            minutes -= max(int((break_end_dt - break_start_dt).total_seconds() // 60), 0)
        return max(minutes, 0)

    def _collect_regular_booking_rows(self, *, start_date: date, end_date: date) -> list[dict[str, Any]]:
        """Нормализует обычные бронирования для отчетов."""
        start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
        end_dt = timezone.make_aware(datetime.combine(end_date, time.max))
        bookings = (
            Booking.objects.filter(
                provider_location_id__in=self.location_ids,
                start_time__gte=start_dt,
                start_time__lte=end_dt,
            )
            .select_related('provider_location', 'employee__user', 'service', 'status')
            .order_by('start_time', 'id')
        )
        rows: list[dict[str, Any]] = []
        for item in bookings:
            rows.append({
                'id': item.id,
                'kind': 'booking',
                'location_id': item.provider_location_id,
                'location_name': item.provider_location.name if item.provider_location_id else '',
                'employee_id': item.employee_id,
                'employee_name': item.employee.user.get_full_name() if item.employee_id and item.employee and item.employee.user else '',
                'service_id': item.service_id,
                'service_name': self._get_service_name(item.service_id),
                'service_path': self._get_service_path(item.service_id),
                'top_category_name': self._get_top_category_name(item.service_id),
                'booking_code': item.code,
                'source': item.source,
                'status': item.status.name if item.status_id and item.status else '',
                'completed_by_actor': item.completed_by_actor or '',
                'start_time': item.start_time,
                'end_time': item.end_time,
                'occupied_duration_minutes': item.occupied_duration_minutes or max(
                    int((item.end_time - item.start_time).total_seconds() // 60),
                    0,
                ),
                'amount': item.price or Decimal('0'),
            })
        return rows

    def _collect_manual_booking_rows(self, *, start_date: date, end_date: date) -> list[dict[str, Any]]:
        """Нормализует manual booking v2 для отчетов."""
        start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
        end_dt = timezone.make_aware(datetime.combine(end_date, time.max))
        bookings = (
            ManualBooking.objects.filter(
                provider_location_id__in=self.location_ids,
                start_time__gte=start_dt,
                start_time__lte=end_dt,
            )
            .select_related('provider_location', 'employee__user', 'service')
            .order_by('start_time', 'id')
        )
        rows: list[dict[str, Any]] = []
        for item in bookings:
            rows.append({
                'id': item.id,
                'kind': 'manual_booking',
                'location_id': item.provider_location_id,
                'location_name': item.provider_location.name if item.provider_location_id else '',
                'employee_id': item.employee_id,
                'employee_name': item.employee.user.get_full_name() if item.employee_id and item.employee and item.employee.user else '',
                'service_id': item.service_id,
                'service_name': self._get_service_name(item.service_id),
                'service_path': self._get_service_path(item.service_id),
                'top_category_name': self._get_top_category_name(item.service_id),
                'booking_code': item.code,
                'source': item.source,
                'status': item.status,
                'completed_by_actor': item.completed_by_actor or '',
                'start_time': item.start_time,
                'end_time': item.end_time,
                'occupied_duration_minutes': item.occupied_duration_minutes or max(
                    int((item.end_time - item.start_time).total_seconds() // 60),
                    0,
                ),
                'amount': item.price or Decimal('0'),
            })
        return rows

    def _collect_all_booking_rows(self, *, start_date: date, end_date: date) -> list[dict[str, Any]]:
        """Объединяет обычные и manual бронирования."""
        rows = self._collect_regular_booking_rows(start_date=start_date, end_date=end_date)
        rows.extend(self._collect_manual_booking_rows(start_date=start_date, end_date=end_date))
        rows.sort(key=lambda item: (item['start_time'], item['id']))
        return rows

    def _build_staff_roster_report(self, *, start_date: date | None, end_date: date | None) -> dict[str, Any]:
        """Формирует отчет по штату филиала/организации."""
        entries = self._get_staff_roster_entries()
        rows = [
            {
                'location_name': item.location_name,
                'employee_name': item.full_name,
                'email': item.email,
                'status': 'roster_active' if item.is_active else 'roster_dismissed',
                'hire_date': item.hire_date.isoformat() if item.hire_date else None,
                'dismissed_at': item.dismissal_datetime.isoformat() if item.dismissal_datetime else None,
                'service_summary': item.service_summary,
                'service_count': item.service_count,
            }
            for item in entries
        ]
        return {
            'title': str(_('Staff roster')),
            'summary': {
                'total_staff': len(entries),
                'active_staff': sum(1 for item in entries if item.is_active),
                'dismissed_staff': sum(1 for item in entries if not item.is_active),
            },
            'rows': rows,
        }

    def _build_services_price_list_report(self, *, start_date: date | None, end_date: date | None) -> dict[str, Any]:
        """Формирует отчет по прайс-листу услуг."""
        rows = []
        queryset = (
            ProviderLocationService.objects.filter(location_id__in=self.location_ids)
            .select_related('location', 'service', 'pet_type')
            .order_by('location__name', 'service__name', 'pet_type__name', 'size_code')
        )
        for item in queryset:
            rows.append({
                'location_name': item.location.name,
                'service_name': self._get_service_name(item.service_id),
                'service_path': self._get_service_path(item.service_id),
                'top_category_name': self._get_top_category_name(item.service_id),
                'pet_type_name': self._get_pet_type_name(item.pet_type_id),
                'size_code': item.size_code,
                'price': str(item.price),
                'duration_minutes': item.duration_minutes,
                'technical_break_minutes': item.tech_break_minutes,
                'is_active': item.is_active,
            })
        return {
            'title': str(_('Services and prices')),
            'summary': {
                'total_rows': len(rows),
                'active_rows': sum(1 for item in rows if item['is_active']),
                'distinct_services': len({item['service_path'] for item in rows}),
            },
            'rows': rows,
        }

    def _build_service_coverage_report(self, *, start_date: date | None, end_date: date | None) -> dict[str, Any]:
        """Формирует отчет по покрытию услуг персоналом."""
        roster_entries = self._get_staff_roster_entries()
        active_staff_keys = {
            (item.employee_id, item.location_id)
            for item in roster_entries
            if item.is_active
        }
        active_services = (
            ProviderLocationService.objects.filter(location_id__in=self.location_ids, is_active=True)
            .select_related('location', 'service')
            .order_by('location__name', 'service__name')
        )
        coverage_counts = defaultdict(set)
        for link in EmployeeLocationService.objects.filter(
            provider_location_id__in=self.location_ids
        ).values('employee_id', 'provider_location_id', 'service_id'):
            key = (link['employee_id'], link['provider_location_id'])
            if key in active_staff_keys:
                coverage_counts[(link['provider_location_id'], link['service_id'])].add(link['employee_id'])

        rows = []
        for item in active_services:
            covered_employees = coverage_counts.get((item.location_id, item.service_id), set())
            rows.append({
                'location_name': item.location.name,
                'service_name': self._get_service_name(item.service_id),
                'service_path': self._get_service_path(item.service_id),
                'top_category_name': self._get_top_category_name(item.service_id),
                'covered_staff_count': len(covered_employees),
                'is_covered': bool(covered_employees),
            })

        uncovered_rows = [item for item in rows if not item['is_covered']]
        category_summary = defaultdict(lambda: {'services': 0, 'covered': 0, 'uncovered': 0})
        for item in rows:
            category = item['top_category_name'] or str(_('Uncategorized'))
            summary = category_summary[category]
            summary['services'] += 1
            if item['is_covered']:
                summary['covered'] += 1
            else:
                summary['uncovered'] += 1

        return {
            'title': str(_('Service coverage')),
            'summary': {
                'total_services': len(rows),
                'covered_services': len(rows) - len(uncovered_rows),
                'uncovered_services': len(uncovered_rows),
            },
            'rows': rows,
            'uncovered_rows': uncovered_rows,
            'categories': [
                {
                    'category_name': key,
                    **value,
                }
                for key, value in sorted(category_summary.items())
            ],
        }

    def _build_staff_schedule_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует матричный отчет по графику персонала."""
        roster_entries = [item for item in self._get_staff_roster_entries() if item.is_active]
        dates = self._iter_dates(start_date, end_date)
        employee_ids = [item.employee_id for item in roster_entries]
        location_ids = [item.location_id for item in roster_entries]

        schedules = (
            Schedule.objects.filter(
                employee_id__in=employee_ids,
                provider_location_id__in=location_ids,
                is_working=True,
            )
            .only(
                'employee_id',
                'provider_location_id',
                'day_of_week',
                'start_time',
                'end_time',
                'break_start',
                'break_end',
            )
        )
        schedule_map = {
            (item.employee_id, item.provider_location_id, item.day_of_week): item
            for item in schedules
        }

        vacation_map: dict[tuple[int, int | None], list[Vacation]] = defaultdict(list)
        for item in Vacation.objects.filter(
            employee_id__in=employee_ids,
            start_date__lte=end_date,
            end_date__gte=start_date,
        ).filter(Q(provider_location_id__in=location_ids) | Q(provider_location__isnull=True)):
            vacation_map[(item.employee_id, item.provider_location_id)].append(item)

        sick_leave_map: dict[tuple[int, int | None], list[SickLeave]] = defaultdict(list)
        for item in SickLeave.objects.filter(
            employee_id__in=employee_ids,
            start_date__lte=end_date,
        ).filter(Q(provider_location_id__in=location_ids) | Q(provider_location__isnull=True)).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=start_date)
        ):
            sick_leave_map[(item.employee_id, item.provider_location_id)].append(item)

        rows = []
        for staff_item in roster_entries:
            cells = []
            for current_date in dates:
                weekday = current_date.weekday()
                schedule = schedule_map.get((staff_item.employee_id, staff_item.location_id, weekday))
                absence_code = ''
                for absence in vacation_map.get((staff_item.employee_id, staff_item.location_id), []) + vacation_map.get((staff_item.employee_id, None), []):
                    if absence.start_date <= current_date <= absence.end_date:
                        absence_code = 'vacation'
                        break
                if not absence_code:
                    for absence in sick_leave_map.get((staff_item.employee_id, staff_item.location_id), []) + sick_leave_map.get((staff_item.employee_id, None), []):
                        if absence.start_date <= current_date and (absence.end_date is None or current_date <= absence.end_date):
                            absence_code = 'sick_leave'
                            break

                if absence_code == 'vacation':
                    label = str(_('Vacation'))
                    cell_type = 'vacation'
                elif absence_code == 'sick_leave':
                    label = str(_('Sick leave'))
                    cell_type = 'sick_leave'
                elif schedule and schedule.start_time and schedule.end_time:
                    label = f'{schedule.start_time.strftime("%H:%M")} - {schedule.end_time.strftime("%H:%M")}'
                    cell_type = 'shift'
                else:
                    label = ''
                    cell_type = 'off'
                cells.append({
                    'date': current_date.isoformat(),
                    'label': label,
                    'type': cell_type,
                })

            rows.append({
                'location_name': staff_item.location_name,
                'employee_name': staff_item.full_name,
                'cells': cells,
            })

        return {
            'title': str(_('Staff schedule')),
            'dates': [item.isoformat() for item in dates],
            'rows': rows,
            'summary': {
                'employees': len(rows),
                'days': len(dates),
            },
        }

    def _build_staff_load_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует отчет по загрузке персонала."""
        roster_entries = [item for item in self._get_staff_roster_entries() if item.is_active]
        if not roster_entries:
            return {
                'title': str(_('Staff load')),
                'summary': {
                    'active_staff': 0,
                    'available_minutes': 0,
                    'busy_minutes': 0,
                    'free_minutes': 0,
                    'utilization_percent': 0,
                },
                'rows': [],
                'locations': [],
            }

        dates = self._iter_dates(start_date, end_date)
        employee_ids = [item.employee_id for item in roster_entries]
        location_ids = [item.location_id for item in roster_entries]

        schedules = Schedule.objects.filter(
            employee_id__in=employee_ids,
            provider_location_id__in=location_ids,
            is_working=True,
        ).only(
            'employee_id',
            'provider_location_id',
            'day_of_week',
            'start_time',
            'end_time',
            'break_start',
            'break_end',
        )
        schedule_map = {
            (item.employee_id, item.provider_location_id, item.day_of_week): item
            for item in schedules
        }

        vacations = defaultdict(list)
        for item in Vacation.objects.filter(
            employee_id__in=employee_ids,
            start_date__lte=end_date,
            end_date__gte=start_date,
        ).filter(Q(provider_location_id__in=location_ids) | Q(provider_location__isnull=True)):
            vacations[(item.employee_id, item.provider_location_id)].append(item)

        sick_leaves = defaultdict(list)
        for item in SickLeave.objects.filter(
            employee_id__in=employee_ids,
            start_date__lte=end_date,
        ).filter(Q(provider_location_id__in=location_ids) | Q(provider_location__isnull=True)).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=start_date)
        ):
            sick_leaves[(item.employee_id, item.provider_location_id)].append(item)

        booking_rows = self._collect_all_booking_rows(start_date=start_date, end_date=end_date)
        busy_minutes_map = defaultdict(int)
        completed_count_map = defaultdict(int)
        revenue_map = defaultdict(Decimal)
        for item in booking_rows:
            if item['employee_id'] is None or item['location_id'] is None:
                continue
            key = (item['employee_id'], item['location_id'])
            if item['status'] != BOOKING_STATUS_CANCELLED:
                busy_minutes_map[key] += int(item['occupied_duration_minutes'] or 0)
            if item['status'] == BOOKING_STATUS_COMPLETED:
                completed_count_map[key] += 1
                revenue_map[key] += Decimal(item['amount'] or 0)

        rows = []
        location_summary = defaultdict(lambda: {
            'active_staff': 0,
            'available_minutes': 0,
            'busy_minutes': 0,
            'free_minutes': 0,
            'completed_bookings': 0,
            'revenue': Decimal('0'),
        })
        for staff_item in roster_entries:
            available_minutes = 0
            absence_days = 0
            for current_date in dates:
                weekday = current_date.weekday()
                is_absent = False
                for absence in vacations.get((staff_item.employee_id, staff_item.location_id), []) + vacations.get((staff_item.employee_id, None), []):
                    if absence.start_date <= current_date <= absence.end_date:
                        is_absent = True
                        break
                if not is_absent:
                    for absence in sick_leaves.get((staff_item.employee_id, staff_item.location_id), []) + sick_leaves.get((staff_item.employee_id, None), []):
                        if absence.start_date <= current_date and (absence.end_date is None or current_date <= absence.end_date):
                            is_absent = True
                            break
                if is_absent:
                    absence_days += 1
                    continue

                schedule = schedule_map.get((staff_item.employee_id, staff_item.location_id, weekday))
                if schedule:
                    available_minutes += self._duration_minutes(
                        schedule.start_time,
                        schedule.end_time,
                        schedule.break_start,
                        schedule.break_end,
                    )

            key = (staff_item.employee_id, staff_item.location_id)
            busy_minutes = busy_minutes_map[key]
            free_minutes = max(available_minutes - busy_minutes, 0)
            utilization_percent = round((busy_minutes / available_minutes) * 100, 2) if available_minutes else 0
            revenue = revenue_map[key]
            row = {
                'location_name': staff_item.location_name,
                'employee_name': staff_item.full_name,
                'available_minutes': available_minutes,
                'busy_minutes': busy_minutes,
                'free_minutes': free_minutes,
                'absence_days': absence_days,
                'completed_bookings': completed_count_map[key],
                'revenue': str(revenue),
                'utilization_percent': utilization_percent,
            }
            rows.append(row)

            summary_row = location_summary[staff_item.location_name]
            summary_row['active_staff'] += 1
            summary_row['available_minutes'] += available_minutes
            summary_row['busy_minutes'] += busy_minutes
            summary_row['free_minutes'] += free_minutes
            summary_row['completed_bookings'] += completed_count_map[key]
            summary_row['revenue'] += revenue

        summary: dict[str, Any] = {
            'active_staff': len(rows),
            'available_minutes': sum(item['available_minutes'] for item in rows),
            'busy_minutes': sum(item['busy_minutes'] for item in rows),
            'free_minutes': sum(item['free_minutes'] for item in rows),
        }
        summary['utilization_percent'] = round(
            (summary['busy_minutes'] / summary['available_minutes']) * 100,
            2,
        ) if summary['available_minutes'] else 0

        return {
            'title': str(_('Staff load')),
            'summary': summary,
            'rows': rows,
            'locations': [
                {
                    'location_name': key,
                    'active_staff': value['active_staff'],
                    'available_minutes': value['available_minutes'],
                    'busy_minutes': value['busy_minutes'],
                    'free_minutes': value['free_minutes'],
                    'completed_bookings': value['completed_bookings'],
                    'revenue': str(value['revenue']),
                    'utilization_percent': round(
                        (value['busy_minutes'] / value['available_minutes']) * 100,
                        2,
                    ) if value['available_minutes'] else 0,
                }
                for key, value in sorted(location_summary.items())
            ],
        }

    def _build_staff_performance_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует отчет по оказанным услугам персоналом."""
        booking_rows = self._collect_all_booking_rows(start_date=start_date, end_date=end_date)
        completed_rows = [item for item in booking_rows if item['status'] == BOOKING_STATUS_COMPLETED]
        load_report = self._build_staff_load_report(start_date=start_date, end_date=end_date)

        per_employee_summary: dict[tuple[str, str], dict[str, Any]] = {}
        per_employee_service: DefaultDict[tuple[str, str, str], dict[str, Any]] = defaultdict(
            lambda: {
                'completed_count': 0,
                'revenue': Decimal('0'),
                'busy_minutes': 0,
                'sources': defaultdict(int),
            },
        )

        load_lookup = {
            (item['location_name'], item['employee_name']): item
            for item in load_report['rows']
        }
        for item in completed_rows:
            employee_name = item['employee_name'] or str(_('Unassigned'))
            location_name = item['location_name'] or ''
            service_name = item['service_name']
            key = (location_name, employee_name)
            summary = per_employee_summary.setdefault(
                key,
                {
                    'location_name': location_name,
                    'employee_name': employee_name,
                    'completed_count': 0,
                    'revenue': Decimal('0'),
                    'busy_minutes': 0,
                    'available_minutes': load_lookup.get(key, {}).get('available_minutes', 0),
                    'free_minutes': load_lookup.get(key, {}).get('free_minutes', 0),
                },
            )
            summary['completed_count'] += 1
            summary['revenue'] += Decimal(item['amount'] or 0)
            summary['busy_minutes'] += int(item['occupied_duration_minutes'] or 0)

            detail_key = (location_name, employee_name, service_name)
            detail = per_employee_service[detail_key]
            detail['completed_count'] += 1
            detail['revenue'] += Decimal(item['amount'] or 0)
            detail['busy_minutes'] += int(item['occupied_duration_minutes'] or 0)
            sources = cast(DefaultDict[str, int], detail['sources'])
            sources[item['source'] or 'unknown'] += 1

        summary_rows = []
        for value in per_employee_summary.values():
            available_minutes = value['available_minutes']
            busy_minutes = value['busy_minutes']
            summary_rows.append({
                'location_name': value['location_name'],
                'employee_name': value['employee_name'],
                'completed_count': value['completed_count'],
                'revenue': str(value['revenue']),
                'busy_minutes': busy_minutes,
                'available_minutes': available_minutes,
                'free_minutes': max(available_minutes - busy_minutes, 0),
                'utilization_percent': round((busy_minutes / available_minutes) * 100, 2) if available_minutes else 0,
            })

        detail_rows = []
        for (location_name, employee_name, service_name), value in sorted(per_employee_service.items()):
            detail_rows.append({
                'location_name': location_name,
                'employee_name': employee_name,
                'service_name': service_name,
                'completed_count': value['completed_count'],
                'revenue': str(value['revenue']),
                'busy_minutes': value['busy_minutes'],
                'booking_service_count': value['sources'].get(Booking.BookingSource.BOOKING_SERVICE, 0),
                'manual_entry_count': value['sources'].get(Booking.BookingSource.MANUAL_ENTRY, 0)
                + value['sources'].get(ManualBooking.Source.MANUAL_ENTRY, 0),
            })

        return {
            'title': str(_('Staff performance')),
            'summary': {
                'employees': len(summary_rows),
                'completed_services': sum(item['completed_count'] for item in summary_rows),
                'revenue': str(sum(Decimal(item['revenue']) for item in summary_rows) if summary_rows else Decimal('0')),
            },
            'employees': summary_rows,
            'details': detail_rows,
        }

    def _build_bookings_summary_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует агрегированный отчет по источникам и статусам записей."""
        booking_rows = self._collect_all_booking_rows(start_date=start_date, end_date=end_date)
        source_summary = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})
        status_summary = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})
        category_summary = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})
        location_summary = defaultdict(lambda: {'count': 0, 'amount': Decimal('0')})

        auto_completed_count = 0
        overdue_not_completed_count = 0
        now = timezone.now()
        for item in booking_rows:
            source_summary[item['source']]['count'] += 1
            source_summary[item['source']]['amount'] += Decimal(item['amount'] or 0)
            status_summary[item['status']]['count'] += 1
            status_summary[item['status']]['amount'] += Decimal(item['amount'] or 0)
            category_key = item['top_category_name'] or str(_('Uncategorized'))
            category_summary[category_key]['count'] += 1
            category_summary[category_key]['amount'] += Decimal(item['amount'] or 0)
            location_summary[item['location_name'] or '']['count'] += 1
            location_summary[item['location_name'] or '']['amount'] += Decimal(item['amount'] or 0)

            if item['kind'] == 'booking' and item['source'] == Booking.BookingSource.BOOKING_SERVICE:
                if item['status'] == BOOKING_STATUS_COMPLETED and item['completed_by_actor'] == COMPLETED_BY_SYSTEM:
                    auto_completed_count += 1
                if item['status'] != BOOKING_STATUS_COMPLETED and item['end_time'] < now:
                    overdue_not_completed_count += 1

        return {
            'title': str(_('Bookings summary')),
            'summary': {
                'total_bookings': len(booking_rows),
                'total_amount': str(sum(Decimal(item['amount'] or 0) for item in booking_rows) if booking_rows else Decimal('0')),
                'auto_completed_count': auto_completed_count,
                'overdue_not_completed_count': overdue_not_completed_count,
            },
            'sources': [
                {
                    'source': key,
                    'count': value['count'],
                    'amount': str(value['amount']),
                }
                for key, value in sorted(source_summary.items())
            ],
            'statuses': [
                {
                    'status': key,
                    'count': value['count'],
                    'amount': str(value['amount']),
                }
                for key, value in sorted(status_summary.items())
            ],
            'categories': [
                {
                    'category_name': key,
                    'count': value['count'],
                    'amount': str(value['amount']),
                }
                for key, value in sorted(category_summary.items())
            ],
            'locations': [
                {
                    'location_name': key,
                    'count': value['count'],
                    'amount': str(value['amount']),
                }
                for key, value in sorted(location_summary.items())
            ],
            'rows': [
                {
                    'location_name': item['location_name'],
                    'employee_name': item['employee_name'],
                    'service_name': item['service_name'],
                    'service_path': item['service_path'],
                    'source': item['source'],
                    'status': item['status'],
                    'start_time': timezone.localtime(item['start_time']).isoformat() if timezone.is_aware(item['start_time']) else item['start_time'].isoformat(),
                    'amount': str(item['amount']),
                }
                for item in booking_rows
            ],
        }

    def _build_financial_revenue_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует финансовый отчет по выручке за период."""
        completed_rows = [
            item
            for item in self._collect_all_booking_rows(start_date=start_date, end_date=end_date)
            if item['status'] == BOOKING_STATUS_COMPLETED
        ]
        total_revenue = sum((Decimal(item['amount'] or 0) for item in completed_rows), Decimal('0'))
        total_bookings = len(completed_rows)

        location_summary = defaultdict(lambda: {'completed_count': 0, 'revenue': Decimal('0')})
        employee_summary = defaultdict(lambda: {'completed_count': 0, 'revenue': Decimal('0')})
        service_summary = defaultdict(lambda: {'completed_count': 0, 'revenue': Decimal('0')})
        detailed_rows = []

        for item in completed_rows:
            amount = Decimal(item['amount'] or 0)
            location_key = item['location_name'] or ''
            employee_key = (item['location_name'] or '', item['employee_name'] or str(_('Unassigned')))
            service_key = (item['location_name'] or '', item['service_name'] or '')

            location_summary[location_key]['completed_count'] += 1
            location_summary[location_key]['revenue'] += amount

            employee_summary[employee_key]['completed_count'] += 1
            employee_summary[employee_key]['revenue'] += amount

            service_summary[service_key]['completed_count'] += 1
            service_summary[service_key]['revenue'] += amount

            detailed_rows.append({
                'booking_code': item['booking_code'],
                'location_name': item['location_name'],
                'employee_name': item['employee_name'] or str(_('Unassigned')),
                'service_name': item['service_name'],
                'start_time': timezone.localtime(item['start_time']).isoformat() if timezone.is_aware(item['start_time']) else item['start_time'].isoformat(),
                'amount': str(amount),
            })

        return {
            'title': str(_('Financial revenue')),
            'summary': {
                'total_bookings': total_bookings,
                'total_revenue': str(total_revenue),
                'average_ticket': str((total_revenue / total_bookings).quantize(Decimal('0.01')) if total_bookings else Decimal('0.00')),
            },
            'locations': [
                {
                    'location_name': key,
                    'completed_count': value['completed_count'],
                    'revenue': str(value['revenue']),
                    'average_ticket': str((value['revenue'] / value['completed_count']).quantize(Decimal('0.01')) if value['completed_count'] else Decimal('0.00')),
                }
                for key, value in sorted(location_summary.items())
            ],
            'employees': [
                {
                    'location_name': location_name,
                    'employee_name': employee_name,
                    'completed_count': value['completed_count'],
                    'revenue': str(value['revenue']),
                    'average_ticket': str((value['revenue'] / value['completed_count']).quantize(Decimal('0.01')) if value['completed_count'] else Decimal('0.00')),
                }
                for (location_name, employee_name), value in sorted(employee_summary.items())
            ],
            'services': [
                {
                    'location_name': location_name,
                    'service_name': service_name,
                    'completed_count': value['completed_count'],
                    'revenue': str(value['revenue']),
                    'average_ticket': str((value['revenue'] / value['completed_count']).quantize(Decimal('0.01')) if value['completed_count'] else Decimal('0.00')),
                }
                for (location_name, service_name), value in sorted(service_summary.items())
            ],
            'rows': detailed_rows,
        }

    def _build_platform_settlement_report(self, *, start_date: date, end_date: date) -> dict[str, Any]:
        """Формирует отчет по взаиморасчетам провайдера с платформой."""
        invoice_queryset = Invoice.objects.filter(provider=self.provider)
        invoice_queryset = invoice_queryset.filter(
            Q(start_date__isnull=True, issued_at__date__gte=start_date, issued_at__date__lte=end_date)
            | Q(start_date__lte=end_date, end_date__gte=start_date)
        ).select_related('currency')

        invoices = list(invoice_queryset.order_by('-issued_at', '-id'))
        invoice_ids = [item.id for item in invoices]
        invoice_lines = list(
            InvoiceLine.objects.filter(invoice_id__in=invoice_ids)
            .select_related(
                'invoice',
                'booking__provider_location',
                'booking__employee__user',
                'booking__service',
            )
            .order_by('invoice__issued_at', 'invoice__number', 'id')
        )

        payments = list(
            PaymentHistory.objects.filter(provider=self.provider).filter(
                Q(due_date__gte=start_date, due_date__lte=end_date)
                | Q(payment_date__gte=start_date, payment_date__lte=end_date)
                | Q(created_at__date__gte=start_date, created_at__date__lte=end_date)
            )
            .select_related('invoice', 'currency')
            .order_by('-due_date', '-id')
        )

        invoice_total_amount = sum((Decimal(item.amount or 0) for item in invoices), Decimal('0'))
        invoice_total_commission = sum((Decimal(item.commission or 0) for item in invoice_lines), Decimal('0'))
        invoice_total_with_vat = sum(
            (
                Decimal(item.total_with_vat or 0)
                if Decimal(item.total_with_vat or 0) > 0
                else Decimal(item.commission or 0) + Decimal(item.vat_amount or 0)
            )
            for item in invoice_lines
        )
        paid_amount = sum((Decimal(item.amount or 0) for item in payments if item.status == 'paid'), Decimal('0'))
        outstanding_amount = sum((Decimal(item.amount or 0) for item in payments if item.status in {'pending', 'overdue'}), Decimal('0'))
        overdue_amount = sum((Decimal(item.amount or 0) for item in payments if item.status == 'overdue'), Decimal('0'))

        location_summary = defaultdict(lambda: {
            'invoice_count': 0,
            'invoiced_commission': Decimal('0'),
            'invoiced_total_with_vat': Decimal('0'),
        })
        line_rows = []
        touched_invoices_by_location: dict[str, set[int]] = defaultdict(set)
        for item in invoice_lines:
            location_name = item.booking.provider_location.name if item.booking.provider_location_id else ''
            touched_invoices_by_location[location_name].add(item.invoice_id)
            total_with_vat = Decimal(item.total_with_vat or 0)
            if total_with_vat <= 0:
                total_with_vat = Decimal(item.commission or 0) + Decimal(item.vat_amount or 0)
            summary = location_summary[location_name]
            summary['invoiced_commission'] += Decimal(item.commission or 0)
            summary['invoiced_total_with_vat'] += total_with_vat

            line_rows.append({
                'invoice_number': item.invoice.number,
                'booking_code': item.booking.code,
                'location_name': location_name,
                'employee_name': item.booking.employee.user.get_full_name() if item.booking.employee_id and item.booking.employee and item.booking.employee.user else str(_('Unassigned')),
                'service_name': self._get_service_name(item.booking.service_id),
                'amount': str(item.amount),
                'commission': str(item.commission),
                'rate': str(item.rate),
                'vat_amount': str(item.vat_amount),
                'total_with_vat': str(total_with_vat),
            })

        for location_name, invoice_set in touched_invoices_by_location.items():
            location_summary[location_name]['invoice_count'] = len(invoice_set)

        return {
            'title': str(_('Platform settlement')),
            'summary': {
                'invoice_count': len(invoices),
                'payment_count': len(payments),
                'invoiced_amount': str(invoice_total_amount),
                'invoiced_commission': str(invoice_total_commission),
                'invoiced_total_with_vat': str(invoice_total_with_vat),
                'paid_amount': str(paid_amount),
                'outstanding_amount': str(outstanding_amount),
                'overdue_amount': str(overdue_amount),
            },
            'locations': [
                {
                    'location_name': key,
                    'invoice_count': value['invoice_count'],
                    'invoiced_commission': str(value['invoiced_commission']),
                    'invoiced_total_with_vat': str(value['invoiced_total_with_vat']),
                }
                for key, value in sorted(location_summary.items())
            ],
            'invoices': [
                {
                    'invoice_number': item.number,
                    'invoice_status': item.status,
                    'start_date': item.start_date.isoformat() if item.start_date else None,
                    'end_date': item.end_date.isoformat() if item.end_date else None,
                    'issued_at': timezone.localtime(item.issued_at).isoformat() if timezone.is_aware(item.issued_at) else item.issued_at.isoformat(),
                    'invoiced_amount': str(item.amount),
                }
                for item in invoices
            ],
            'payments': [
                {
                    'invoice_number': item.invoice.number if item.invoice_id else '',
                    'status': item.status,
                    'due_date': item.due_date.isoformat(),
                    'payment_date': item.payment_date.isoformat() if item.payment_date else None,
                    'paid_amount': str(item.amount),
                    'description': item.description,
                }
                for item in payments
            ],
            'rows': line_rows,
        }

    def _build_workbook(self, report_payload: dict[str, Any]) -> openpyxl.Workbook:
        """Создает XLSX-книгу из JSON payload отчета."""
        workbook = openpyxl.Workbook()
        first_sheet = workbook.active
        assert first_sheet is not None
        workbook.remove(first_sheet)

        report_code = report_payload['report_code']
        data = report_payload['data']
        summary_rows = [{'metric': key, 'value': value} for key, value in data.get('summary', {}).items()]
        self._append_sheet(workbook, 'Summary', summary_rows)

        if report_code == 'staff_roster':
            self._append_sheet(workbook, 'Staff', data.get('rows', []))
        elif report_code == 'services_price_list':
            self._append_sheet(workbook, 'Services', data.get('rows', []))
        elif report_code == 'service_coverage':
            self._append_sheet(workbook, 'Coverage', data.get('rows', []))
            self._append_sheet(workbook, 'Uncovered', data.get('uncovered_rows', []))
            self._append_sheet(workbook, 'Categories', data.get('categories', []))
        elif report_code == 'staff_schedule':
            self._append_staff_schedule_sheet(workbook, data)
        elif report_code == 'staff_load':
            self._append_sheet(workbook, 'Employees', data.get('rows', []))
            self._append_sheet(workbook, 'Locations', data.get('locations', []))
        elif report_code == 'staff_performance':
            self._append_sheet(workbook, 'Employees', data.get('employees', []))
            self._append_sheet(workbook, 'Details', data.get('details', []))
        elif report_code == 'bookings_summary':
            self._append_sheet(workbook, 'Sources', data.get('sources', []))
            self._append_sheet(workbook, 'Statuses', data.get('statuses', []))
            self._append_sheet(workbook, 'Categories', data.get('categories', []))
            self._append_sheet(workbook, 'Locations', data.get('locations', []))
            self._append_sheet(workbook, 'Bookings', data.get('rows', []))
        elif report_code == 'financial_revenue':
            self._append_sheet(workbook, 'Locations', data.get('locations', []))
            self._append_sheet(workbook, 'Employees', data.get('employees', []))
            self._append_sheet(workbook, 'Services', data.get('services', []))
            self._append_sheet(workbook, 'Bookings', data.get('rows', []))
        elif report_code == 'platform_settlement':
            self._append_sheet(workbook, 'Locations', data.get('locations', []))
            self._append_sheet(workbook, 'Invoices', data.get('invoices', []))
            self._append_sheet(workbook, 'Payments', data.get('payments', []))
            self._append_sheet(workbook, 'Details', data.get('rows', []))
        else:
            self._append_sheet(workbook, 'Data', data.get('rows', []))

        return workbook

    def _sheet_title(self, key: str) -> str:
        if self.language_code == 'ru':
            return self.RUSSIAN_SHEET_TITLES.get(key, key)
        return str(self.SHEET_TITLES.get(key, key))

    def _chunk_sequence(self, items: list[Any], chunk_size: int) -> list[list[Any]]:
        """Разбивает последовательность на чанки фиксированного размера."""
        if chunk_size <= 0:
            return [items]
        return [items[index:index + chunk_size] for index in range(0, len(items), chunk_size)] or [[]]

    def _paginated_sheet_title(
        self,
        title: str,
        *,
        row_part: int | None = None,
        column_part: int | None = None,
    ) -> str:
        """Строит имя листа с номером части в пределах лимита Excel."""
        suffix_parts: list[str] = []
        if row_part is not None:
            suffix_parts.append(f'r{row_part}')
        if column_part is not None:
            suffix_parts.append(f'c{column_part}')
        suffix = f" ({'-'.join(suffix_parts)})" if suffix_parts else ""
        trimmed_title = title[: max(1, 31 - len(suffix))]
        return f"{trimmed_title}{suffix}"[:31]

    def _style_header_cell(self, cell: openpyxl.cell.cell.Cell) -> None:
        """Применяет общий стиль к ячейке заголовка."""
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center')

    def _field_label(self, key: str) -> str:
        if self.language_code == 'ru':
            return self.RUSSIAN_FIELD_LABELS.get(key, key.replace('_', ' ').title())
        return str(self.FIELD_LABELS.get(key, key.replace('_', ' ').title()))

    def _humanize_value(self, value: str) -> str:
        return value.replace('_', ' ').replace('-', ' ').title()

    def _coerce_to_date(self, value: Any) -> date | None:
        if value is None or value == '':
            return None
        if isinstance(value, datetime):
            aware = timezone.make_aware(value) if timezone.is_naive(value) else value
            return timezone.localtime(aware).date()
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, str):
            head = value[:10]
            parsed = parse_date(head)
            if parsed is not None:
                return parsed
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            except ValueError:
                return None
            aware = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
            return timezone.localtime(aware).date()
        return None

    def _format_localized_date(self, value: Any) -> Any:
        coerced = self._coerce_to_date(value)
        if coerced is None:
            return value
        with translation.override(self.language_code):
            return date_format(coerced, format='SHORT_DATE_FORMAT', use_l10n=True)

    def _staff_roster_status_label(self, value: str) -> str:
        labels = {
            'ru': ('Активен', 'Уволен'),
            'de': ('Aktiv', 'Entlassen'),
            'me': ('Aktivan', 'Otpušten'),
            'en': ('Active', 'Dismissed'),
        }
        active, dismissed = labels.get(self.language_code, labels['en'])
        return active if value == 'roster_active' else dismissed

    def _format_export_value(self, field_name: str, value: Any) -> Any:
        if isinstance(value, (list, tuple)):
            return ', '.join(str(item) for item in value)
        if field_name == 'metric' and isinstance(value, str):
            return self._field_label(value)
        if field_name in {'hire_date', 'dismissed_at'}:
            return self._format_localized_date(value)
        if isinstance(value, bool):
            if self.language_code == 'ru':
                return 'Да' if value else 'Нет'
            return str(_('Yes')) if value else str(_('No'))
        if field_name == 'source' and isinstance(value, str):
            if self.language_code == 'ru':
                source_labels: dict[str, str] = {
                    Booking.BookingSource.BOOKING_SERVICE.value: 'Через сервис',
                    Booking.BookingSource.MANUAL_ENTRY.value: 'Ручной ввод',
                }
                return source_labels.get(value, self._humanize_value(value))
            source_labels_en: dict[str, Any] = {
                Booking.BookingSource.BOOKING_SERVICE.value: _('Online booking'),
                Booking.BookingSource.MANUAL_ENTRY.value: _('Manual entry'),
            }
            return str(source_labels_en.get(value, self._humanize_value(value)))
        if field_name == 'status' and isinstance(value, str):
            if value in ('roster_active', 'roster_dismissed'):
                return self._staff_roster_status_label(value)
            if self.language_code == 'ru':
                status_labels = {
                    BOOKING_STATUS_ACTIVE: 'Активна',
                    BOOKING_STATUS_CANCELLED: 'Отменена',
                    BOOKING_STATUS_COMPLETED: 'Завершена',
                }
                return status_labels.get(value, self._humanize_value(value))
            status_labels = {
                BOOKING_STATUS_ACTIVE: _('Active'),
                BOOKING_STATUS_CANCELLED: _('Cancelled'),
                BOOKING_STATUS_COMPLETED: _('Completed'),
            }
            return str(status_labels.get(value, self._humanize_value(value)))
        if field_name == 'completed_by_actor' and isinstance(value, str):
            if self.language_code == 'ru':
                actor_labels = {
                    COMPLETED_BY_SYSTEM: 'Система',
                    'provider': 'Провайдер',
                    'client': 'Клиент',
                }
                return actor_labels.get(value, self._humanize_value(value))
            actor_labels = {
                COMPLETED_BY_SYSTEM: _('System'),
                'provider': _('Provider'),
                'client': _('Client'),
            }
            return str(actor_labels.get(value, self._humanize_value(value)))
        return value

    def _append_sheet(self, workbook: openpyxl.Workbook, title: str, rows: list[dict[str, Any]]) -> None:
        """Добавляет стандартный лист-таблицу."""
        base_title = self._sheet_title(title)[:31] or str(_('Sheet'))
        if not rows:
            worksheet = workbook.create_sheet(title=base_title)
            worksheet['A1'] = str(_('No data'))
            return

        columns = list(rows[0].keys())
        row_chunk_size = max(1, self.EXCEL_MAX_ROWS - 1)
        row_chunks = self._chunk_sequence(rows, row_chunk_size)
        column_chunks = self._chunk_sequence(columns, self.EXCEL_MAX_COLUMNS)

        for row_chunk_index, row_chunk in enumerate(row_chunks, start=1):
            for column_chunk_index, column_chunk in enumerate(column_chunks, start=1):
                worksheet = workbook.create_sheet(
                    title=self._paginated_sheet_title(
                        base_title,
                        row_part=row_chunk_index if len(row_chunks) > 1 else None,
                        column_part=column_chunk_index if len(column_chunks) > 1 else None,
                    )
                )
                for excel_column_index, column_name in enumerate(column_chunk, start=1):
                    cell = worksheet.cell(row=1, column=excel_column_index, value=self._field_label(column_name))
                    self._style_header_cell(cell)

                for excel_row_index, row_data in enumerate(row_chunk, start=2):
                    for excel_column_index, column_name in enumerate(column_chunk, start=1):
                        value = self._format_export_value(column_name, row_data.get(column_name))
                        worksheet.cell(row=excel_row_index, column=excel_column_index, value=value)

                self._autofit_columns(worksheet)

    def _append_staff_schedule_sheet(self, workbook: openpyxl.Workbook, data: dict[str, Any]) -> None:
        """Добавляет матричный лист для графика персонала."""
        base_title = self._sheet_title('Schedule')[:31] or str(_('Sheet'))
        dates = data.get('dates', [])
        rows = data.get('rows', [])
        if not rows:
            worksheet = workbook.create_sheet(title=base_title)
            worksheet['A1'] = str(_('No data'))
            return

        row_chunk_size = max(1, self.EXCEL_MAX_ROWS - 1)
        row_chunks = self._chunk_sequence(rows, row_chunk_size)
        fixed_header_count = 2
        date_chunk_size = max(1, self.EXCEL_MAX_COLUMNS - fixed_header_count)
        date_chunks = self._chunk_sequence(dates, date_chunk_size)

        for row_chunk_index, row_chunk in enumerate(row_chunks, start=1):
            for date_chunk_index, date_chunk in enumerate(date_chunks, start=1):
                worksheet = workbook.create_sheet(
                    title=self._paginated_sheet_title(
                        base_title,
                        row_part=row_chunk_index if len(row_chunks) > 1 else None,
                        column_part=date_chunk_index if len(date_chunks) > 1 else None,
                    )
                )
                headers = ['location_name', 'employee_name', *date_chunk]
                date_offset = (date_chunk_index - 1) * date_chunk_size
                for column_index, header in enumerate(headers, start=1):
                    header_value = self._field_label(header) if header in {'location_name', 'employee_name'} else header
                    cell = worksheet.cell(row=1, column=column_index, value=header_value)
                    self._style_header_cell(cell)

                for row_index, row_data in enumerate(row_chunk, start=2):
                    worksheet.cell(row=row_index, column=1, value=row_data.get('location_name'))
                    worksheet.cell(row=row_index, column=2, value=row_data.get('employee_name'))
                    date_cells = row_data.get('cells', [])[date_offset:date_offset + len(date_chunk)]
                    for column_index, cell_data in enumerate(date_cells, start=3):
                        worksheet.cell(row=row_index, column=column_index, value=cell_data.get('label'))

                self._autofit_columns(worksheet)

    def _autofit_columns(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Подбирает адекватную ширину колонок."""
        for column_cells in worksheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            if not values:
                continue
            max_width = min(max(len(value) for value in values) + 2, 48)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_width
