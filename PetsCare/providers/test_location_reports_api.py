"""Тесты operational reports для страницы филиала."""

import shutil
import tempfile
from datetime import timedelta
from decimal import Decimal
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework.test import APIClient

from booking.constants import BOOKING_STATUS_COMPLETED, COMPLETED_BY_USER, COMPLETION_REASON_MANUAL
from booking.models import Booking, BookingStatus
from catalog.models import Service
from geolocation.models import Address
from pets.models import Pet, PetType
from providers.models import (
    Employee,
    EmployeeLocationRole,
    EmployeeLocationService,
    EmployeeProvider,
    Provider,
    ProviderLocation,
    ProviderLocationService,
    ProviderReportExportJob,
)
from providers.reporting_services import ProviderLocationReportingService
from providers.tasks import generate_provider_report_export_task
from scheduling.models import Vacation

User = get_user_model()


class ProviderLocationReportsAPITests(TestCase):
    """Проверяет staff payload и report endpoint филиала."""

    def setUp(self):
        self.client = APIClient()
        self.today = timezone.localdate()

        self.provider = Provider.objects.create(
            name='Report Vet',
            phone_number='+38267010101',
            email='report-vet@example.com',
            activation_status='active',
            is_active=True,
        )
        self.address = Address.objects.create(
            country='Montenegro',
            city='Podgorica',
            street='Reports street',
            house_number='7',
            formatted_address='Reports street 7',
            latitude=42.44,
            longitude=19.26,
            validation_status='valid',
        )
        self.location = ProviderLocation.objects.create(
            provider=self.provider,
            name='Central reports branch',
            structured_address=self.address,
            phone_number='+38267010102',
            email='branch@example.com',
            is_active=True,
        )

        self.admin_user = User.objects.create_user(
            email='owner@example.com',
            password='password123',
            username='owner@example.com',
            phone_number='+38267010103',
        )
        self.admin_user.add_role('provider_admin')
        self.admin_employee = Employee.objects.create(user=self.admin_user, is_active=True)
        EmployeeProvider.objects.create(
            employee=self.admin_employee,
            provider=self.provider,
            role=EmployeeProvider.ROLE_OWNER,
            is_owner=True,
            start_date=self.today,
        )
        self.client.force_authenticate(self.admin_user)

        self.grooming_service = Service.objects.create(
            code='report_grooming',
            name='Grooming',
            name_en='Grooming',
            level=0,
        )
        self.pet_type = PetType.objects.create(name='Dog', code='dog')
        self.consultation_service = Service.objects.create(
            code='report_consultation',
            name='Consultation',
            name_en='Consultation',
            level=0,
        )
        ProviderLocationService.objects.create(
            location=self.location,
            service=self.grooming_service,
            pet_type=self.pet_type,
            size_code='S',
            price=Decimal('20.00'),
            duration_minutes=30,
            is_active=True,
        )
        ProviderLocationService.objects.create(
            location=self.location,
            service=self.consultation_service,
            pet_type=self.pet_type,
            size_code='S',
            price=Decimal('25.00'),
            duration_minutes=45,
            is_active=True,
        )

        self.active_employee = self._create_employee(
            email='active@example.com',
            phone='+38267010104',
            first_name='Active',
            last_name='Worker',
            start_offset_days=20,
            is_active=True,
        )
        EmployeeLocationService.objects.create(
            employee=self.active_employee,
            provider_location=self.location,
            service=self.grooming_service,
        )

        self.dismissed_employee = self._create_employee(
            email='dismissed@example.com',
            phone='+38267010105',
            first_name='Dismissed',
            last_name='Worker',
            start_offset_days=40,
            is_active=False,
            end_offset_days=2,
        )
        Vacation.objects.create(
            employee=self.active_employee,
            provider_location=self.location,
            start_date=self.today,
            end_date=self.today + timedelta(days=1),
            is_approved=True,
        )

    def _create_employee(
        self,
        *,
        email: str,
        phone: str,
        first_name: str,
        last_name: str,
        start_offset_days: int,
        is_active: bool,
        end_offset_days: int | None = None,
    ) -> Employee:
        """Создает сотрудника и связь с локацией."""
        user = User.objects.create_user(
            email=email,
            password='password123',
            username=email,
            first_name=first_name,
            last_name=last_name,
            phone_number=phone,
        )
        employee = Employee.objects.create(user=user, is_active=True)
        EmployeeProvider.objects.create(
            employee=employee,
            provider=self.provider,
            role=EmployeeProvider.ROLE_SERVICE_WORKER,
            start_date=self.today - timedelta(days=start_offset_days),
            end_date=self.today - timedelta(days=end_offset_days) if end_offset_days is not None else None,
        )
        employee.locations.add(self.location)
        EmployeeLocationRole.objects.create(
            employee=employee,
            provider_location=self.location,
            role=EmployeeLocationRole.ROLE_SERVICE_WORKER,
            is_active=is_active,
            end_date=timezone.now() - timedelta(days=end_offset_days) if end_offset_days is not None else None,
        )
        return employee

    def test_staff_payload_contains_hire_date_dismissed_at_and_services(self):
        url = reverse('providers:location-staff-list', kwargs={'pk': self.location.id})

        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        employees = response.data['employees']  # type: ignore[index]
        active_row = next(item for item in employees if item['id'] == self.active_employee.id)
        dismissed_row = next(item for item in employees if item['id'] == self.dismissed_employee.id)

        self.assertIn('hire_date', active_row)
        self.assertEqual(active_row['service_count'], 1)
        self.assertEqual(active_row['service_names'], ['Grooming'])
        self.assertTrue(active_row['service_summary'])
        self.assertFalse(dismissed_row['is_active'])
        self.assertIsNotNone(dismissed_row['dismissed_at'])

    def test_service_coverage_report_returns_json_and_xlsx(self):
        json_url = reverse('providers:location-reports', kwargs={'location_pk': self.location.id})

        response = self.client.get(
            json_url,
            {'report': 'service_coverage', 'scope': 'location', 'output': 'json'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['report_code'], 'service_coverage')  # type: ignore[index]
        self.assertEqual(response.data['data']['summary']['uncovered_services'], 1)  # type: ignore[index]

        xlsx_response = self.client.get(
            json_url,
            {'report': 'service_coverage', 'scope': 'location', 'output': 'xlsx'},
        )
        self.assertEqual(xlsx_response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            xlsx_response['Content-Type'],
        )

    def test_standard_xlsx_export_splits_large_tables_into_multiple_sheets(self):
        service = ProviderLocationReportingService(
            location=self.location,
            scope='location',
            language_code='en',
        )
        service.EXCEL_MAX_ROWS = 4
        service.EXCEL_MAX_COLUMNS = 2
        workbook = openpyxl.Workbook()
        active_sheet = workbook.active
        assert active_sheet is not None
        workbook.remove(active_sheet)

        service._append_sheet(
            workbook,
            'Data',
            [
                {'alpha': 'A1', 'beta': 'B1', 'gamma': 'C1'},
                {'alpha': 'A2', 'beta': 'B2', 'gamma': 'C2'},
                {'alpha': 'A3', 'beta': 'B3', 'gamma': 'C3'},
                {'alpha': 'A4', 'beta': 'B4', 'gamma': 'C4'},
            ],
        )

        self.assertEqual(workbook.sheetnames, ['Data (r1-c1)', 'Data (r1-c2)', 'Data (r2-c1)', 'Data (r2-c2)'])
        self.assertEqual(workbook['Data (r1-c1)'].max_row, 4)
        self.assertEqual(workbook['Data (r1-c1)'].max_column, 2)
        self.assertEqual(workbook['Data (r1-c1)']['A2'].value, 'A1')
        self.assertEqual(workbook['Data (r1-c2)']['A1'].value, 'Gamma')
        self.assertEqual(workbook['Data (r2-c2)']['A2'].value, 'C4')

    def test_staff_schedule_xlsx_export_splits_rows_and_date_columns(self):
        service = ProviderLocationReportingService(
            location=self.location,
            scope='location',
            language_code='en',
        )
        service.EXCEL_MAX_ROWS = 4
        service.EXCEL_MAX_COLUMNS = 4
        workbook = openpyxl.Workbook()
        active_sheet = workbook.active
        assert active_sheet is not None
        workbook.remove(active_sheet)

        service._append_staff_schedule_sheet(
            workbook,
            {
                'dates': ['2026-03-20', '2026-03-21', '2026-03-22'],
                'rows': [
                    {
                        'location_name': 'Central reports branch',
                        'employee_name': f'Employee {index}',
                        'cells': [
                            {'label': f'D{day_index}-{index}'}
                            for day_index in range(1, 4)
                        ],
                    }
                    for index in range(1, 5)
                ],
            },
        )

        self.assertEqual(workbook.sheetnames, ['Schedule (r1-c1)', 'Schedule (r1-c2)', 'Schedule (r2-c1)', 'Schedule (r2-c2)'])
        self.assertEqual(workbook['Schedule (r1-c1)']['A1'].value, 'Branch')
        self.assertEqual(workbook['Schedule (r1-c1)']['C1'].value, '2026-03-20')
        self.assertEqual(workbook['Schedule (r1-c2)']['C1'].value, '2026-03-22')
        self.assertEqual(workbook['Schedule (r2-c2)']['B2'].value, 'Employee 4')
        self.assertEqual(workbook['Schedule (r2-c2)']['C2'].value, 'D3-4')

    def test_provider_reports_allows_organization_scope_for_owner(self):
        second_location = ProviderLocation.objects.create(
            provider=self.provider,
            name='North branch',
            structured_address=self.address,
            phone_number='+38267010109',
            email='north@example.com',
            is_active=True,
        )
        ProviderLocationService.objects.create(
            location=second_location,
            service=self.grooming_service,
            pet_type=self.pet_type,
            size_code='S',
            price=Decimal('18.00'),
            duration_minutes=25,
            is_active=True,
        )

        response = self.client.get(
            reverse('providers:provider-reports', kwargs={'provider_id': self.provider.id}),
            {
                'report': 'service_coverage',
                'scope': 'provider',
                'location_id': self.location.id,
                'output': 'json',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['report_code'], 'service_coverage')  # type: ignore[index]
        self.assertEqual(response.data['scope'], 'provider')  # type: ignore[index]
        self.assertTrue(response.data['context']['organization_scope_allowed'])  # type: ignore[index]
        self.assertEqual(response.data['context']['location_id'], self.location.id)  # type: ignore[index]
        self.assertGreaterEqual(response.data['data']['summary']['total_services'], 2)  # type: ignore[index]

    def test_report_access_filters_locations_for_branch_staff_and_blocks_provider_scope(self):
        second_location = ProviderLocation.objects.create(
            provider=self.provider,
            name='South branch',
            structured_address=self.address,
            phone_number='+38267010110',
            email='south@example.com',
            is_active=True,
        )
        branch_user = User.objects.create_user(
            email='branch@example.com',
            password='password123',
            username='branch@example.com',
            phone_number='+38267010111',
        )
        branch_employee = Employee.objects.create(user=branch_user, is_active=True)
        EmployeeProvider.objects.create(
            employee=branch_employee,
            provider=self.provider,
            role=EmployeeProvider.ROLE_SERVICE_WORKER,
            start_date=self.today,
        )
        EmployeeLocationRole.objects.create(
            employee=branch_employee,
            provider_location=self.location,
            role=EmployeeLocationRole.ROLE_SERVICE_WORKER,
            is_active=True,
        )

        self.client.force_authenticate(branch_user)

        locations_response = self.client.get(
            reverse('providers:provider-location-list-create'),
            {'provider': self.provider.id, 'report_access': 1},
        )

        self.assertEqual(locations_response.status_code, 200)
        self.assertEqual([item['id'] for item in locations_response.data], [self.location.id])  # type: ignore[index]
        self.assertNotIn(second_location.id, [item['id'] for item in locations_response.data])  # type: ignore[index]

        location_report_response = self.client.get(
            reverse('providers:provider-reports', kwargs={'provider_id': self.provider.id}),
            {
                'report': 'service_coverage',
                'scope': 'location',
                'location_id': self.location.id,
                'output': 'json',
            },
        )
        self.assertEqual(location_report_response.status_code, 200)
        self.assertFalse(location_report_response.data['context']['organization_scope_allowed'])  # type: ignore[index]

        provider_scope_response = self.client.get(
            reverse('providers:provider-reports', kwargs={'provider_id': self.provider.id}),
            {
                'report': 'service_coverage',
                'scope': 'provider',
                'location_id': self.location.id,
                'output': 'json',
            },
        )
        self.assertEqual(provider_scope_response.status_code, 403)

    def test_branch_staff_can_read_assigned_location_directory_and_services_only_for_own_branch(self):
        second_location = ProviderLocation.objects.create(
            provider=self.provider,
            name='South branch',
            structured_address=self.address,
            phone_number='+38267010120',
            email='south-services@example.com',
            is_active=True,
        )
        ProviderLocationService.objects.create(
            location=second_location,
            service=self.grooming_service,
            pet_type=self.pet_type,
            size_code='S',
            price=Decimal('30.00'),
            duration_minutes=30,
            is_active=True,
        )
        self.location.served_pet_types.add(self.pet_type)
        second_location.served_pet_types.add(self.pet_type)

        branch_employee = self._create_employee(
            email='branch-services@example.com',
            phone='+38267010121',
            first_name='Branch',
            last_name='Worker',
            start_offset_days=0,
            is_active=True,
        )

        self.client.force_authenticate(branch_employee.user)

        locations_response = self.client.get(
            reverse('providers:provider-location-list-create'),
            {'provider': self.provider.id},
        )

        self.assertEqual(locations_response.status_code, 200)
        self.assertEqual([item['id'] for item in locations_response.data], [self.location.id])  # type: ignore[index]

        services_response = self.client.get(
            reverse('providers:provider-location-service-list-create'),
            {'location': self.location.id},
        )
        self.assertEqual(services_response.status_code, 200)
        self.assertEqual(len(services_response.data), 2)  # type: ignore[arg-type]

        own_matrix_response = self.client.get(
            reverse('providers:location-price-matrix', kwargs={'pk': self.location.id}),
        )
        self.assertEqual(own_matrix_response.status_code, 200)
        self.assertEqual(len(own_matrix_response.data), 2)  # type: ignore[arg-type]

        foreign_matrix_response = self.client.get(
            reverse('providers:location-price-matrix', kwargs={'pk': second_location.id}),
        )
        self.assertEqual(foreign_matrix_response.status_code, 403)

    def test_financial_revenue_report_returns_expected_aggregates(self):
        customer = User.objects.create_user(
            email='customer@example.com',
            password='password123',
            username='customer@example.com',
            phone_number='+38267010112',
        )
        pet = Pet.objects.create(
            name='Rex',
            pet_type=self.pet_type,
            weight=Decimal('8.50'),
        )
        completed_status = BookingStatus.objects.get_or_create(
            name=BOOKING_STATUS_COMPLETED,
            defaults={'code': BOOKING_STATUS_COMPLETED, 'label': 'Completed'},
        )[0]
        Booking.objects.create(
            user=customer,
            escort_owner=customer,
            pet=pet,
            provider=self.provider,
            provider_location=self.location,
            employee=self.active_employee,
            service=self.grooming_service,
            status=completed_status,
            start_time=timezone.now() - timedelta(days=1, hours=2),
            end_time=timezone.now() - timedelta(days=1, hours=1, minutes=30),
            occupied_duration_minutes=30,
            price=Decimal('42.00'),
            completed_at=timezone.now() - timedelta(days=1, hours=1, minutes=20),
            completed_by_actor=COMPLETED_BY_USER,
            completed_by_user=self.admin_user,
            completion_reason_code=COMPLETION_REASON_MANUAL,
            code='REVTEST01',
        )

        response = self.client.get(
            reverse('providers:provider-reports', kwargs={'provider_id': self.provider.id}),
            {
                'report': 'financial_revenue',
                'scope': 'provider',
                'location_id': self.location.id,
                'start_date': (self.today - timedelta(days=7)).isoformat(),
                'end_date': self.today.isoformat(),
                'output': 'json',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['report_code'], 'financial_revenue')  # type: ignore[index]
        self.assertEqual(response.data['data']['summary']['total_bookings'], 1)  # type: ignore[index]
        self.assertEqual(response.data['data']['summary']['total_revenue'], '42.00')  # type: ignore[index]
        self.assertEqual(response.data['data']['summary']['average_ticket'], '42.00')  # type: ignore[index]
        self.assertEqual(response.data['data']['locations'][0]['location_name'], self.location.name)  # type: ignore[index]

    def test_branch_staff_export_list_hides_provider_scope_jobs(self):
        branch_user = User.objects.create_user(
            email='branch-jobs@example.com',
            password='password123',
            username='branch-jobs@example.com',
            phone_number='+38267010113',
        )
        branch_employee = Employee.objects.create(user=branch_user, is_active=True)
        EmployeeProvider.objects.create(
            employee=branch_employee,
            provider=self.provider,
            role=EmployeeProvider.ROLE_SERVICE_WORKER,
            start_date=self.today,
        )
        EmployeeLocationRole.objects.create(
            employee=branch_employee,
            provider_location=self.location,
            role=EmployeeLocationRole.ROLE_SERVICE_WORKER,
            is_active=True,
        )

        visible_job = ProviderReportExportJob.objects.create(
            provider=self.provider,
            location=self.location,
            requested_by=branch_user,
            report_code='service_coverage',
            scope='location',
            export_format=ProviderReportExportJob.FORMAT_XLSX,
            language_code='en',
            filename='visible.xlsx',
        )
        ProviderReportExportJob.objects.create(
            provider=self.provider,
            location=self.location,
            requested_by=branch_user,
            report_code='bookings_summary',
            scope='provider',
            export_format=ProviderReportExportJob.FORMAT_XLSX,
            language_code='en',
            filename='hidden.xlsx',
        )

        self.client.force_authenticate(branch_user)
        response = self.client.get(
            reverse('providers:provider-report-export-list-create', kwargs={'provider_id': self.provider.id}),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([item['id'] for item in response.data['results']], [visible_job.id])  # type: ignore[index]

    def test_provider_report_export_job_can_be_created_and_downloaded(self):
        temp_media = tempfile.mkdtemp()
        try:
            with self.settings(MEDIA_ROOT=temp_media):
                create_url = reverse(
                    'providers:provider-report-export-list-create',
                    kwargs={'provider_id': self.provider.id},
                )
                with patch('providers.api_views.generate_provider_report_export_task.delay') as delay_mock:
                    response = self.client.post(
                        create_url,
                        {
                            'report': 'service_coverage',
                            'scope': 'location',
                            'location_id': self.location.id,
                            'language': 'en',
                            'export_format': 'xlsx',
                        },
                        format='json',
                    )

                self.assertEqual(response.status_code, 202)
                job_id = response.data['id']  # type: ignore[index]
                delay_mock.assert_called_once_with(job_id)

                generate_provider_report_export_task.run(job_id)

                detail_response = self.client.get(
                    reverse(
                        'providers:provider-report-export-detail',
                        kwargs={'provider_id': self.provider.id, 'job_id': job_id},
                    )
                )
                self.assertEqual(detail_response.status_code, 200)
                self.assertEqual(
                    detail_response.data['status'],
                    ProviderReportExportJob.STATUS_COMPLETED,
                    detail_response.data,  # type: ignore[arg-type]
                )
                self.assertTrue(detail_response.data['download_url'])  # type: ignore[index]

                download_response = self.client.get(
                    reverse(
                        'providers:provider-report-export-download',
                        kwargs={'provider_id': self.provider.id, 'job_id': job_id},
                    )
                )
                self.assertEqual(download_response.status_code, 200)
                self.assertIn(
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    download_response['Content-Type'],
                )
                file_stream = getattr(download_response, 'file_to_stream', None)
                if file_stream is not None:
                    file_stream.close()
                job = ProviderReportExportJob.objects.get(pk=job_id)
                self.assertIsNotNone(job.downloaded_at)
        finally:
            shutil.rmtree(temp_media, ignore_errors=True)

    def test_stale_pending_export_job_is_marked_failed_in_list(self):
        stale_job = ProviderReportExportJob.objects.create(
            provider=self.provider,
            location=self.location,
            requested_by=self.admin_user,
            report_code='service_coverage',
            scope='location',
            export_format=ProviderReportExportJob.FORMAT_XLSX,
            language_code='en',
            filename='stale.xlsx',
        )
        ProviderReportExportJob.objects.filter(pk=stale_job.pk).update(
            created_at=timezone.now() - timedelta(minutes=10),
        )

        response = self.client.get(
            reverse('providers:provider-report-export-list-create', kwargs={'provider_id': self.provider.id}),
        )

        self.assertEqual(response.status_code, 200)
        stale_job.refresh_from_db()
        self.assertEqual(stale_job.status, ProviderReportExportJob.STATUS_FAILED)
        self.assertEqual(stale_job.error_message, 'Export job timed out before processing started.')
        self.assertEqual(response.data['results'][0]['status'], ProviderReportExportJob.STATUS_FAILED)  # type: ignore[index]
