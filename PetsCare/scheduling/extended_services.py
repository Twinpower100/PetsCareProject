"""
Extended services for the scheduling module.

Этот модуль содержит расширенные сервисы для автоматического планирования расписания
с интеграцией системы бронирования и генерацией отчетов.

Основные компоненты:
1. ExtendedSchedulePlannerService - расширенный сервис планирования
2. BookingIntegrationService - интеграция с системой бронирования
3. AdvancedReportGenerator - расширенная генерация отчетов
4. EmergencyRescheduler - экстренное перепланирование при больничных

Особенности реализации:
- Интеграция с существующими бронированиями
- Обработка экстренных ситуаций
- Детальная отчетность
- Экспорт в Excel
"""

from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from .models import (
    Workplace, WorkplaceAllowedServices, ServicePriority,
    Vacation, SickLeave, DayOff, EmployeeSchedule, StaffingRequirement
)
from providers.models import Provider, Employee, Schedule, ProviderSchedule
from booking.models import Booking, TimeSlot
from catalog.models import Service
from .services import SchedulePlannerService, AvailabilityChecker, ConflictResolver


class BookingIntegrationService:
    """
    Сервис для интеграции планирования с системой бронирования.
    
    Основные функции:
    - Проверка существующих бронирований
    - Обновление временных слотов
    - Синхронизация с планированием
    """
    
    def __init__(self, provider: Provider):
        """
        Инициализирует сервис интеграции с бронированиями.
        
        Args:
            provider: Учреждение
        """
        self.provider = provider
    
    def check_existing_bookings(self, start_date: date, end_date: date) -> Dict:
        """
        Проверяет существующие бронирования в периоде.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            
        Returns:
            Dict: Информация о бронированиях
        """
        bookings = Booking.objects.filter(
            provider=self.provider,
            start_time__date__gte=start_date,
            start_time__date__lte=end_date,
            status__name__in=['active', 'pending_confirmation']
        ).select_related('employee', 'service', 'status')
        
        booking_data = {}
        for booking in bookings:
            date_key = booking.start_time.date().isoformat()
            if date_key not in booking_data:
                booking_data[date_key] = []
            
            booking_data[date_key].append({
                'id': booking.id,
                'employee': booking.employee,
                'service': booking.service,
                'start_time': booking.start_time,
                'end_time': booking.end_time,
                'status': booking.status.name
            })
        
        return {
            'total_bookings': bookings.count(),
            'bookings_by_date': booking_data
        }
    
    def update_time_slots(self, schedule_result: Dict) -> Dict:
        """
        Обновляет временные слоты на основе планирования.
        
        Args:
            schedule_result: Результат планирования
            
        Returns:
            Dict: Результат обновления
        """
        updated_slots = []
        created_slots = []
        errors = []
        
        schedule = schedule_result.get('schedule', {})
        
        for date_str, assignments in schedule.items():
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            for assignment in assignments:
                try:
                    # Создаем или обновляем временной слот
                    slot, created = TimeSlot.objects.get_or_create(
                        employee=assignment['employee'],
                        provider=self.provider,
                        start_time=datetime.combine(date_obj, assignment['start_time']),
                        end_time=datetime.combine(date_obj, assignment['end_time']),
                        defaults={'is_available': True}
                    )
                    
                    if created:
                        created_slots.append(slot)
                    else:
                        slot.is_available = True
                        slot.save()
                        updated_slots.append(slot)
                        
                except Exception as e:
                    errors.append({
                        'assignment': assignment,
                        'error': str(e)
                    })
        
        return {
            'updated_slots': len(updated_slots),
            'created_slots': len(created_slots),
            'errors': errors
        }
    
    def find_conflicting_bookings(self, schedule_result: Dict) -> List[Dict]:
        """
        Находит конфликтующие бронирования.
        
        Args:
            schedule_result: Результат планирования
            
        Returns:
            List[Dict]: Список конфликтов
        """
        conflicts = []
        schedule = schedule_result.get('schedule', {})
        
        for date_str, assignments in schedule.items():
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            for assignment in assignments:
                # Проверяем конфликтующие бронирования
                conflicting_bookings = Booking.objects.filter(
                    employee=assignment['employee'],
                    start_time__date=date_obj,
                    start_time__lt=datetime.combine(date_obj, assignment['end_time']),
                    end_time__gt=datetime.combine(date_obj, assignment['start_time']),
                    status__name__in=['active', 'pending_confirmation']
                )
                
                for booking in conflicting_bookings:
                    conflicts.append({
                        'assignment': assignment,
                        'booking': booking,
                        'conflict_type': 'time_overlap',
                        'date': date_obj
                    })
        
        return conflicts


class EmergencyRescheduler:
    """
    Сервис для экстренного перепланирования при больничных.
    
    Основные функции:
    - Обработка внезапных больничных
    - Перепланирование затронутых дней
    - Уведомления о изменениях
    """
    
    def __init__(self, provider: Provider):
        """
        Инициализирует сервис экстренного перепланирования.
        
        Args:
            provider: Учреждение
        """
        self.provider = provider
        self.planner = SchedulePlannerService(provider)
        self.booking_integration = BookingIntegrationService(provider)
    
    def handle_sick_leave(self, employee: Employee, start_date: date, 
                         end_date: date = None) -> Dict:
        """
        Обрабатывает больничный сотрудника.
        
        Args:
            employee: Сотрудник
            start_date: Дата начала больничного
            end_date: Дата окончания больничного (если известна)
            
        Returns:
            Dict: Результат обработки
        """
        if end_date is None:
            end_date = start_date + timedelta(days=7)  # По умолчанию неделя
        
        # Получаем затронутые бронирования
        affected_bookings = Booking.objects.filter(
            employee=employee,
            provider=self.provider,
            start_time__date__gte=start_date,
            start_time__date__lte=end_date,
            status__name__in=['active', 'pending_confirmation']
        )
        
        # Находим альтернативных сотрудников
        alternative_assignments = []
        unassigned_bookings = []
        
        for booking in affected_bookings:
            alternative_employee = self._find_alternative_employee(
                booking.service, booking.start_time.date()
            )
            
            if alternative_employee:
                alternative_assignments.append({
                    'original_booking': booking,
                    'alternative_employee': alternative_employee
                })
            else:
                unassigned_bookings.append(booking)
        
        # Перепланируем затронутые дни
        reschedule_result = self._reschedule_affected_days(
            start_date, end_date, alternative_assignments
        )
        
        return {
            'affected_bookings': len(affected_bookings),
            'alternative_assignments': len(alternative_assignments),
            'unassigned_bookings': len(unassigned_bookings),
            'reschedule_result': reschedule_result,
            'recommendations': self._generate_emergency_recommendations(
                affected_bookings, alternative_assignments, unassigned_bookings
            )
        }
    
    def _find_alternative_employee(self, service: Service, target_date: date) -> Optional[Employee]:
        """
        Находит альтернативного сотрудника для услуги.
        
        Args:
            service: Услуга
            target_date: Дата
            
        Returns:
            Optional[Employee]: Альтернативный сотрудник или None
        """
        available_employees = Employee.objects.filter(
            providers=self.provider,
            services=service,
            is_active=True
        ).exclude(
            # Исключаем сотрудников с больничными
            sickleaves__start_date__lte=target_date,
            sickleaves__end_date__gte=target_date,
            sickleaves__is_confirmed=True
        )
        
        # Проверяем доступность
        for employee in available_employees:
            if self._is_employee_available(employee, target_date):
                return employee
        
        return None
    
    def _is_employee_available(self, employee: Employee, target_date: date) -> bool:
        """
        Проверяет доступность сотрудника.
        
        Args:
            employee: Сотрудник
            target_date: Дата
            
        Returns:
            bool: True если доступен
        """
        # Проверяем расписание
        day_of_week = target_date.weekday()
        try:
            schedule = employee.schedules.get(day_of_week=day_of_week)
            if not schedule.is_working:
                return False
        except Schedule.DoesNotExist:
            return False
        
        # Проверяем существующие бронирования
        existing_bookings = Booking.objects.filter(
            employee=employee,
            start_time__date=target_date,
            status__name__in=['active', 'pending_confirmation']
        )
        
        # Если у сотрудника уже много бронирований, считаем его недоступным
        if existing_bookings.count() >= 8:  # Максимум 8 часов в день
            return False
        
        return True
    
    def _reschedule_affected_days(self, start_date: date, end_date: date,
                                alternative_assignments: List[Dict]) -> Dict:
        """
        Перепланирует затронутые дни.
        
        Args:
            start_date: Дата начала
            end_date: Дата окончания
            alternative_assignments: Альтернативные назначения
            
        Returns:
            Dict: Результат перепланирования
        """
        # Создаем временные слоты для альтернативных сотрудников
        created_slots = []
        
        for assignment in alternative_assignments:
            booking = assignment['original_booking']
            employee = assignment['alternative_employee']
            
            try:
                slot = TimeSlot.objects.create(
                    employee=employee,
                    provider=self.provider,
                    start_time=booking.start_time,
                    end_time=booking.end_time,
                    is_available=True
                )
                created_slots.append(slot)
            except Exception as e:
                pass
        
        return {
            'created_slots': len(created_slots),
            'affected_days': (end_date - start_date).days + 1
        }
    
    def _generate_emergency_recommendations(self, affected_bookings: List[Booking],
                                          alternative_assignments: List[Dict],
                                          unassigned_bookings: List[Booking]) -> List[str]:
        """
        Генерирует рекомендации для экстренной ситуации.
        
        Args:
            affected_bookings: Затронутые бронирования
            alternative_assignments: Альтернативные назначения
            unassigned_bookings: Нераспределенные бронирования
            
        Returns:
            List[str]: Список рекомендаций
        """
        recommendations = []
        
        if unassigned_bookings:
            recommendations.append(
                _('{} bookings could not be reassigned. Consider contacting clients to reschedule.')
                .format(len(unassigned_bookings))
            )
        
        if alternative_assignments:
            recommendations.append(
                _('Successfully reassigned {} bookings to alternative employees.')
                .format(len(alternative_assignments))
            )
        
        recommendations.append(
            _('Consider hiring temporary staff or adjusting schedules to prevent future conflicts.')
        )
        
        return recommendations


class AdvancedReportGenerator:
    """
    Расширенный генератор отчетов.
    
    Основные функции:
    - Детальные отчеты по планированию
    - Анализ эффективности
    - Прогнозирование потребностей
    """
    
    def __init__(self, provider: Provider):
        """
        Инициализирует расширенный генератор отчетов.
        
        Args:
            provider: Учреждение
        """
        self.provider = provider
    
    def generate_comprehensive_report(self, start_date: date, end_date: date,
                                    schedule_result: Dict) -> Dict:
        """
        Генерирует комплексный отчет по планированию.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            schedule_result: Результат планирования
            
        Returns:
            Dict: Комплексный отчет
        """
        # Базовые отчеты
        coverage_report = self._generate_coverage_report(start_date, end_date, schedule_result)
        efficiency_report = self._generate_efficiency_report(start_date, end_date, schedule_result)
        financial_report = self._generate_financial_report(start_date, end_date, schedule_result)
        
        return {
            'provider': self.provider.name,
            'period': _('{} - {}').format(start_date, end_date),
            'generated_at': timezone.now(),
            'coverage': coverage_report,
            'efficiency': efficiency_report,
            'financial': financial_report,
            'summary': self._generate_summary(coverage_report, efficiency_report, financial_report)
        }
    
    def _generate_coverage_report(self, start_date: date, end_date: date,
                                schedule_result: Dict) -> Dict:
        """
        Генерирует отчет по покрытию потребностей.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            schedule_result: Результат планирования
            
        Returns:
            Dict: Отчет по покрытию
        """
        requirements = StaffingRequirement.objects.filter(
            provider=self.provider,
            is_active=True
        )
        
        coverage_data = {}
        total_required = 0
        total_assigned = 0
        
        for requirement in requirements:
            service = requirement.service
            required_count = requirement.required_count
            day_of_week = requirement.day_of_week
            
            # Подсчитываем дни в периоде
            period_days = self._count_weekdays_in_period(start_date, end_date, day_of_week)
            total_required_hours = period_days * required_count * 8
            
            # Подсчитываем назначенные часы
            assigned_hours = self._count_assigned_hours_for_service(
                schedule_result, service, start_date, end_date
            )
            
            coverage_data[service.name] = {
                'required_hours': total_required_hours,
                'assigned_hours': assigned_hours,
                'coverage_percentage': (assigned_hours / total_required_hours * 100) if total_required_hours > 0 else 0,
                'period_days': period_days,
                'daily_requirement': required_count
            }
            
            total_required += total_required_hours
            total_assigned += assigned_hours
        
        return {
            'total_required_hours': total_required,
            'total_assigned_hours': total_assigned,
            'overall_coverage_percentage': (total_assigned / total_required * 100) if total_required > 0 else 0,
            'service_coverage': coverage_data
        }
    
    def _generate_efficiency_report(self, start_date: date, end_date: date,
                                  schedule_result: Dict) -> Dict:
        """
        Генерирует отчет по эффективности планирования.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            schedule_result: Результат планирования
            
        Returns:
            Dict: Отчет по эффективности
        """
        # Анализируем распределение нагрузки
        employee_workload = {}
        workplace_utilization = {}
        
        schedule = schedule_result.get('schedule', {})
        
        for date_str, assignments in schedule.items():
            for assignment in assignments:
                employee = assignment['employee']
                workplace = assignment.get('workplace')
                
                # Подсчитываем нагрузку на сотрудника
                if employee.id not in employee_workload:
                    employee_workload[employee.id] = {
                        'employee': employee,
                        'total_hours': 0,
                        'days_worked': 0,
                        'services': set()
                    }
                
                hours = self._calculate_assignment_hours(assignment)
                employee_workload[employee.id]['total_hours'] += hours
                employee_workload[employee.id]['days_worked'] += 1
                employee_workload[employee.id]['services'].add(assignment['service'].name)
                
                # Подсчитываем использование рабочего места
                if workplace:
                    if workplace.id not in workplace_utilization:
                        workplace_utilization[workplace.id] = {
                            'workplace': workplace,
                            'total_hours': 0,
                            'days_used': 0
                        }
                    
                    workplace_utilization[workplace.id]['total_hours'] += hours
                    workplace_utilization[workplace.id]['days_used'] += 1
        
        return {
            'employee_workload': employee_workload,
            'workplace_utilization': workplace_utilization,
            'average_workload_per_employee': sum(w['total_hours'] for w in employee_workload.values()) / len(employee_workload) if employee_workload else 0,
            'average_workplace_utilization': sum(w['total_hours'] for w in workplace_utilization.values()) / len(workplace_utilization) if workplace_utilization else 0
        }
    
    def _generate_financial_report(self, start_date: date, end_date: date,
                                 schedule_result: Dict) -> Dict:
        """
        Генерирует финансовый отчет.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            schedule_result: Результат планирования
            
        Returns:
            Dict: Финансовый отчет
        """
        # Подсчитываем потенциальный доход
        potential_revenue = 0
        service_revenue = {}
        
        schedule = schedule_result.get('schedule', {})
        
        for date_str, assignments in schedule.items():
            for assignment in assignments:
                service = assignment['service']
                
                # Получаем цену услуги
                try:
                    provider_service = self.provider.provider_services.get(service=service)
                    price = provider_service.price
                except:
                    price = 0
                
                hours = self._calculate_assignment_hours(assignment)
                revenue = price * (hours / 8)  # Предполагаем 8-часовой рабочий день
                
                potential_revenue += revenue
                
                if service.name not in service_revenue:
                    service_revenue[service.name] = 0
                service_revenue[service.name] += revenue
        
        return {
            'potential_revenue': potential_revenue,
            'service_revenue': service_revenue,
            'average_daily_revenue': potential_revenue / ((end_date - start_date).days + 1) if start_date != end_date else potential_revenue
        }
    
    def _count_weekdays_in_period(self, start_date: date, end_date: date, 
                                 weekday: int) -> int:
        """
        Подсчитывает количество определенных дней недели в периоде.
        
        Args:
            start_date: Дата начала периода
            end_date: Дата окончания периода
            weekday: День недели (0-6)
            
        Returns:
            int: Количество дней
        """
        count = 0
        current_date = start_date
        
        while current_date <= end_date:
            if current_date.weekday() == weekday:
                count += 1
            current_date += timedelta(days=1)
        
        return count
    
    def _count_assigned_hours_for_service(self, schedule_result: Dict, service: Service,
                                        start_date: date, end_date: date) -> int:
        """
        Подсчитывает назначенные часы для услуги.
        
        Args:
            schedule_result: Результат планирования
            service: Услуга
            start_date: Дата начала периода
            end_date: Дата окончания периода
            
        Returns:
            int: Количество часов
        """
        hours = 0
        schedule = schedule_result.get('schedule', {})
        
        for date_str, assignments in schedule.items():
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            if start_date <= date_obj <= end_date:
                for assignment in assignments:
                    if assignment['service'] == service:
                        hours += self._calculate_assignment_hours(assignment)
        
        return hours
    
    def _calculate_assignment_hours(self, assignment: Dict) -> int:
        """
        Подсчитывает часы назначения.
        
        Args:
            assignment: Назначение
            
        Returns:
            int: Количество часов
        """
        start_time = assignment.get('start_time')
        end_time = assignment.get('end_time')
        
        if start_time and end_time:
            start_dt = datetime.combine(date.today(), start_time)
            end_dt = datetime.combine(date.today(), end_time)
            duration = end_dt - start_dt
            return int(duration.total_seconds() / 3600)
        
        return 8  # По умолчанию 8 часов
    
    def _generate_summary(self, coverage_report: Dict, efficiency_report: Dict,
                         financial_report: Dict) -> Dict:
        """
        Генерирует сводку отчетов.
        
        Args:
            coverage_report: Отчет по покрытию
            efficiency_report: Отчет по эффективности
            financial_report: Финансовый отчет
            
        Returns:
            Dict: Сводка
        """
        return {
            'overall_coverage': coverage_report.get('overall_coverage_percentage', 0),
            'average_workload': efficiency_report.get('average_workload_per_employee', 0),
            'potential_revenue': financial_report.get('potential_revenue', 0),
            'recommendations': self._generate_report_recommendations(
                coverage_report, efficiency_report, financial_report
            )
        }
    
    def _generate_report_recommendations(self, coverage_report: Dict,
                                       efficiency_report: Dict,
                                       financial_report: Dict) -> List[str]:
        """
        Генерирует рекомендации на основе отчетов.
        
        Args:
            coverage_report: Отчет по покрытию
            efficiency_report: Отчет по эффективности
            financial_report: Финансовый отчет
            
        Returns:
            List[str]: Список рекомендаций
        """
        recommendations = []
        
        # Рекомендации по покрытию
        coverage = coverage_report.get('overall_coverage_percentage', 0)
        if coverage < 80:
            recommendations.append(
                _('Low coverage rate ({:.1f}%). Consider hiring additional staff or adjusting requirements.')
                .format(coverage)
            )
        elif coverage > 95:
            recommendations.append(
                _('Very high coverage rate ({:.1f}%). Consider reducing staff or increasing service offerings.')
                .format(coverage)
            )
        
        # Рекомендации по нагрузке
        avg_workload = efficiency_report.get('average_workload_per_employee', 0)
        if avg_workload < 30:
            recommendations.append(
                _('Low average workload ({:.1f} hours). Consider reducing staff or increasing service demand.')
                .format(avg_workload)
            )
        elif avg_workload > 45:
            recommendations.append(
                _('High average workload ({:.1f} hours). Consider hiring additional staff to prevent burnout.')
                .format(avg_workload)
            )
        
        # Рекомендации по доходам
        revenue = financial_report.get('potential_revenue', 0)
        if revenue > 0:
            recommendations.append(
                _('Potential revenue: ${:,.2f}. Focus on converting potential to actual bookings.')
                .format(revenue)
            )
        
        return recommendations


class ExtendedSchedulePlannerService(SchedulePlannerService):
    """
    Расширенный сервис планирования расписания.
    
    Основные функции:
    - Все функции базового сервиса
    - Интеграция с системой бронирования
    - Экстренное перепланирование
    - Расширенная отчетность
    """
    
    def __init__(self, provider: Provider):
        """
        Инициализирует расширенный сервис планирования.
        
        Args:
            provider: Учреждение
        """
        super().__init__(provider)
        self.booking_integration = BookingIntegrationService(provider)
        self.emergency_rescheduler = EmergencyRescheduler(provider)
        self.advanced_reporter = AdvancedReportGenerator(provider)
    
    def plan_schedule_with_bookings(self, start_date: date, end_date: date,
                                  optimize_preferences: bool = True) -> Dict[str, Any]:
        """
        Планирует расписание с учетом существующих бронирований.
        
        Args:
            start_date: Дата начала планирования
            end_date: Дата окончания планирования
            optimize_preferences: Оптимизировать по предпочтениям
            
        Returns:
            Dict: Результат планирования
        """
        # Проверяем существующие бронирования
        booking_info = self.booking_integration.check_existing_bookings(start_date, end_date)
        
        # Планируем расписание
        schedule_result = self.plan_schedule(start_date, end_date, optimize_preferences)
        
        if not schedule_result.get('success', False):
            return schedule_result
        
        # Проверяем конфликты с бронированиями
        booking_conflicts = self.booking_integration.find_conflicting_bookings(schedule_result)
        schedule_result['booking_conflicts'] = booking_conflicts
        
        # Обновляем временные слоты
        slot_update_result = self.booking_integration.update_time_slots(schedule_result)
        schedule_result['slot_updates'] = slot_update_result
        
        # Генерируем расширенный отчет
        comprehensive_report = self.advanced_reporter.generate_comprehensive_report(
            start_date, end_date, schedule_result
        )
        schedule_result['comprehensive_report'] = comprehensive_report
        
        return schedule_result
    
    def handle_emergency_situation(self, employee: Employee, start_date: date,
                                 end_date: date = None) -> Dict[str, Any]:
        """
        Обрабатывает экстренную ситуацию (больничный, отгул).
        
        Args:
            employee: Сотрудник
            start_date: Дата начала
            end_date: Дата окончания
            
        Returns:
            Dict: Результат обработки
        """
        return self.emergency_rescheduler.handle_sick_leave(employee, start_date, end_date)
    
    def export_comprehensive_report(self, schedule_result: Dict,
                                  filename: str = None) -> bytes:
        """
        Экспортирует комплексный отчет в Excel.
        
        Args:
            schedule_result: Результат планирования
            filename: Имя файла
            
        Returns:
            bytes: Содержимое Excel файла
        """
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        
        # Лист с расписанием
        schedule_ws = wb.active
        schedule_ws.title = _('Schedule')
        
        # Заголовки расписания
        schedule_headers = [
            _('Date'), _('Day of Week'), _('Employee'), 
            _('Service'), _('Workplace'), _('Start Time'), _('End Time'), _('Hours')
        ]
        schedule_ws.append(schedule_headers)
        
        # Данные расписания
        schedule = schedule_result.get('schedule', {})
        for date_str, assignments in schedule.items():
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            for assignment in assignments:
                hours = self._calculate_assignment_hours(assignment)
                row = [
                    date_obj.strftime('%Y-%m-%d'),
                    date_obj.strftime('%A'),
                    str(assignment['employee']),
                    assignment['service'].name,
                    assignment['workplace'].name if assignment.get('workplace') else _('Any'),
                    assignment['start_time'].strftime('%H:%M'),
                    assignment['end_time'].strftime('%H:%M'),
                    hours
                ]
                schedule_ws.append(row)
        
        # Лист с отчетами
        if 'comprehensive_report' in schedule_result:
            report = schedule_result['comprehensive_report']
            
            # Отчет по покрытию
            coverage_ws = wb.create_sheet(_('Coverage Report'))
            coverage_ws.append([_('Service'), _('Required Hours'), _('Assigned Hours'), _('Coverage %')])
            
            for service_name, data in report.get('coverage', {}).get('service_coverage', {}).items():
                coverage_ws.append([
                    service_name,
                    data['required_hours'],
                    data['assigned_hours'],
                    f"{data['coverage_percentage']:.1f}%"
                ])
            
            # Отчет по эффективности
            efficiency_ws = wb.create_sheet(_('Efficiency Report'))
            efficiency_ws.append([_('Employee'), _('Total Hours'), _('Days Worked'), _('Services')])
            
            for employee_data in report.get('efficiency', {}).get('employee_workload', {}).values():
                efficiency_ws.append([
                    str(employee_data['employee']),
                    employee_data['total_hours'],
                    employee_data['days_worked'],
                    ', '.join(employee_data['services'])
                ])
            
            # Финансовый отчет
            financial_ws = wb.create_sheet(_('Financial Report'))
            financial_ws.append([_('Service'), _('Revenue')])
            
            for service_name, revenue in report.get('financial', {}).get('service_revenue', {}).items():
                financial_ws.append([service_name, f"${revenue:,.2f}"])
        
        # Сохраняем в байты
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _calculate_assignment_hours(self, assignment: Dict) -> int:
        """
        Подсчитывает часы назначения.
        
        Args:
            assignment: Назначение
            
        Returns:
            int: Количество часов
        """
        start_time = assignment.get('start_time')
        end_time = assignment.get('end_time')
        
        if start_time and end_time:
            start_dt = datetime.combine(date.today(), start_time)
            end_dt = datetime.combine(date.today(), end_time)
            duration = end_dt - start_dt
            return int(duration.total_seconds() / 3600)
        
        return 8  # По умолчанию 8 часов 