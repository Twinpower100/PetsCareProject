"""
API views для системы отчетности PetCare.

Этот модуль содержит API endpoints для:
1. Генерации отчетов по доходам
2. Генерации отчетов по загруженности сотрудников
3. Генерации отчетов по дебиторской задолженности
4. Генерации отчетов по активности учреждений
5. Генерации отчетов по платежам
6. Генерации отчетов по отменам бронирований
7. Экспорта отчетов в Excel
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from datetime import datetime, timedelta
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .services import (
    IncomeReportService,
    EmployeeWorkloadReportService,
    DebtReportService,
    ActivityReportService,
    PaymentReportService,
    CancellationReportService
)
from providers.models import Provider
from users.models import User


def get_date_range_from_request(request) -> tuple:
    """Извлекает диапазон дат из запроса."""
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    else:
        start_date = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        end_date = datetime.now()
    
    return start_date, end_date


def get_providers_from_request(request) -> List[Provider]:
    """Извлекает список провайдеров из запроса."""
    provider_ids = request.GET.getlist('providers')
    if provider_ids:
        return Provider.objects.filter(id__in=provider_ids)
    return None


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def income_report(request):
    """
    API endpoint для генерации отчета по доходам.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = IncomeReportService(request.user)
        report_data = service.generate_income_report(start_date, end_date, providers)
        
        # Проверяем формат ответа
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_income_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_workload_report(request):
    """
    API endpoint для генерации отчета по загруженности сотрудников.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = EmployeeWorkloadReportService(request.user)
        report_data = service.generate_workload_report(start_date, end_date, providers)
        
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_workload_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def debt_report(request):
    """
    API endpoint для генерации отчета по дебиторской задолженности.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = DebtReportService(request.user)
        report_data = service.generate_debt_report(start_date, end_date, providers)
        
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_debt_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def activity_report(request):
    """
    API endpoint для генерации отчета по активности учреждений.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = ActivityReportService(request.user)
        report_data = service.generate_activity_report(start_date, end_date, providers)
        
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_activity_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def payment_report(request):
    """
    API endpoint для генерации отчета по платежам.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = PaymentReportService(request.user)
        report_data = service.generate_payment_report(start_date, end_date, providers)
        
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_payment_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cancellation_report(request):
    """
    API endpoint для генерации отчета по отменам бронирований.
    
    Query parameters:
    - start_date: Начальная дата (YYYY-MM-DD)
    - end_date: Конечная дата (YYYY-MM-DD)
    - providers: Список ID провайдеров (может быть несколько)
    - format: Формат ответа (json/excel)
    
    Returns:
    - JSON с данными отчета или Excel файл
    """
    try:
        start_date, end_date = get_date_range_from_request(request)
        providers = get_providers_from_request(request)
        
        service = CancellationReportService(request.user)
        report_data = service.generate_cancellation_report(start_date, end_date, providers)
        
        response_format = request.GET.get('format', 'json')
        
        if response_format.lower() == 'excel':
            return generate_cancellation_excel_report(report_data)
        
        return Response({
            'success': True,
            'data': report_data
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)


# Функции для генерации Excel отчетов

def generate_income_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по доходам."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    # Заголовки
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Данные
    summary_data = [
        ['Общий доход', f"{report_data['summary']['total_income']:.2f}"],
        ['Количество бронирований', report_data['summary']['total_bookings']],
        ['Общая комиссия', f"{report_data['summary']['total_commission']:.2f}"],
        ['Средний чек', f"{report_data['summary']['average_booking_value']:.2f}"],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по провайдерам
    ws_providers = wb.create_sheet("По провайдерам")
    headers = ['Провайдер', 'Доход', 'Количество бронирований']
    for col, header in enumerate(headers, 1):
        cell = ws_providers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, provider in enumerate(report_data['by_provider'], 2):
        ws_providers.cell(row=row, column=1, value=provider['provider__name'])
        ws_providers.cell(row=row, column=2, value=float(provider['income']))
        ws_providers.cell(row=row, column=3, value=provider['bookings_count'])
    
    # Лист по услугам
    ws_services = wb.create_sheet("По услугам")
    headers = ['Услуга', 'Доход', 'Количество бронирований']
    for col, header in enumerate(headers, 1):
        cell = ws_services.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, service in enumerate(report_data['by_service'], 2):
        ws_services.cell(row=row, column=1, value=service['service__name'])
        ws_services.cell(row=row, column=2, value=float(service['income']))
        ws_services.cell(row=row, column=3, value=service['bookings_count'])
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_providers, ws_services]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=income_report.xlsx'
    wb.save(response)
    return response


def generate_workload_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по загруженности сотрудников."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    summary_data = [
        ['Общее количество часов', f"{report_data['summary']['total_hours']:.2f}"],
        ['Общее количество бронирований', report_data['summary']['total_bookings']],
        ['Количество сотрудников', report_data['summary']['total_employees']],
        ['Средняя эффективность', f"{report_data['summary']['average_efficiency']:.2f}"],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по сотрудникам
    ws_employees = wb.create_sheet("По сотрудникам")
    headers = ['Имя', 'Фамилия', 'Email', 'Провайдер', 'Часы', 'Бронирования', 'Доход', 'Эффективность']
    for col, header in enumerate(headers, 1):
        cell = ws_employees.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, employee in enumerate(report_data['by_employee'], 2):
        ws_employees.cell(row=row, column=1, value=employee['employee__user__first_name'])
        ws_employees.cell(row=row, column=2, value=employee['employee__user__last_name'])
        ws_employees.cell(row=row, column=3, value=employee['employee__user__email'])
        ws_employees.cell(row=row, column=4, value=employee['provider__name'])
        ws_employees.cell(row=row, column=5, value=float(employee['total_hours'] or 0))
        ws_employees.cell(row=row, column=6, value=employee['bookings_count'])
        ws_employees.cell(row=row, column=7, value=float(employee['total_income'] or 0))
        ws_employees.cell(row=row, column=8, value=float(employee['efficiency'] or 0))
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_employees]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=workload_report.xlsx'
    wb.save(response)
    return response


def generate_debt_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по дебиторской задолженности."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    summary_data = [
        ['Общая задолженность', f"{report_data['summary']['total_debt']:.2f}"],
        ['Количество провайдеров с задолженностью', report_data['summary']['providers_with_debt']],
        ['Средняя задолженность', f"{report_data['summary']['average_debt']:.2f}"],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по провайдерам
    ws_providers = wb.create_sheet("По провайдерам")
    headers = ['Провайдер', 'Задолженность', 'Дни просрочки', 'Номер договора', 'Статус договора']
    for col, header in enumerate(headers, 1):
        cell = ws_providers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, provider in enumerate(report_data['providers'], 2):
        ws_providers.cell(row=row, column=1, value=provider['provider_name'])
        ws_providers.cell(row=row, column=2, value=float(provider['total_debt']))
        ws_providers.cell(row=row, column=3, value=provider['overdue_days'])
        ws_providers.cell(row=row, column=4, value=provider['contract_number'])
        ws_providers.cell(row=row, column=5, value=provider['contract_status'])
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_providers]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=debt_report.xlsx'
    wb.save(response)
    return response


def generate_activity_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по активности учреждений."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    summary_data = [
        ['Количество провайдеров', report_data['summary']['total_providers']],
        ['Общее количество бронирований', report_data['summary']['total_bookings']],
        ['Общий доход', f"{report_data['summary']['total_income']:.2f}"],
        ['Процент завершения', f"{report_data['summary']['completion_rate']:.2f}%"],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по провайдерам
    ws_providers = wb.create_sheet("По провайдерам")
    headers = ['Провайдер', 'Всего бронирований', 'Завершенных', 'Отмененных', 'Доход', 'Услуг', 'Клиентов']
    for col, header in enumerate(headers, 1):
        cell = ws_providers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, provider in enumerate(report_data['by_provider'], 2):
        ws_providers.cell(row=row, column=1, value=provider['provider__name'])
        ws_providers.cell(row=row, column=2, value=provider['total_bookings'])
        ws_providers.cell(row=row, column=3, value=provider['completed_bookings'])
        ws_providers.cell(row=row, column=4, value=provider['cancelled_bookings'])
        ws_providers.cell(row=row, column=5, value=float(provider['total_income'] or 0))
        ws_providers.cell(row=row, column=6, value=provider['unique_services'])
        ws_providers.cell(row=row, column=7, value=provider['unique_customers'])
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_providers]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=activity_report.xlsx'
    wb.save(response)
    return response


def generate_payment_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по платежам."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    summary_data = [
        ['Получено', f"{report_data['summary']['total_received']:.2f}"],
        ['Ожидается', f"{report_data['summary']['total_expected']:.2f}"],
        ['Количество платежей', report_data['summary']['total_payments']],
        ['Процент успешности', f"{report_data['summary']['success_rate']:.2f}%"],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по провайдерам
    ws_providers = wb.create_sheet("По провайдерам")
    headers = ['Провайдер', 'Получено', 'Ожидается', 'Количество', 'Процент успешности']
    for col, header in enumerate(headers, 1):
        cell = ws_providers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, provider in enumerate(report_data['by_provider'], 2):
        ws_providers.cell(row=row, column=1, value=provider['booking__provider__name'])
        ws_providers.cell(row=row, column=2, value=float(provider['total_received'] or 0))
        ws_providers.cell(row=row, column=3, value=float(provider['total_expected'] or 0))
        ws_providers.cell(row=row, column=4, value=provider['payment_count'])
        ws_providers.cell(row=row, column=5, value=float(provider['success_rate'] or 0))
    
    # Лист просроченных платежей
    ws_overdue = wb.create_sheet("Просроченные платежи")
    headers = ['Номер счета', 'Провайдер', 'Сумма', 'Валюта', 'Дата выставления', 'Дни просрочки']
    for col, header in enumerate(headers, 1):
        cell = ws_overdue.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, payment in enumerate(report_data['overdue_payments'], 2):
        ws_overdue.cell(row=row, column=1, value=payment['invoice_number'])
        ws_overdue.cell(row=row, column=2, value=payment['provider_name'])
        ws_overdue.cell(row=row, column=3, value=float(payment['amount']))
        ws_overdue.cell(row=row, column=4, value=payment['currency'])
        ws_overdue.cell(row=row, column=5, value=payment['issued_at'].strftime('%Y-%m-%d'))
        ws_overdue.cell(row=row, column=6, value=payment['overdue_days'])
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_providers, ws_overdue]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payment_report.xlsx'
    wb.save(response)
    return response


def generate_cancellation_excel_report(report_data: Dict[str, Any]) -> HttpResponse:
    """Генерирует Excel отчет по отменам бронирований."""
    wb = openpyxl.Workbook()
    
    # Лист с общей статистикой
    ws_summary = wb.active
    ws_summary.title = "Общая статистика"
    
    headers = ['Показатель', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    summary_data = [
        ['Общее количество отмен', report_data['summary']['total_cancellations']],
        ['Отмены клиентами', report_data['summary']['client_cancellations']],
        ['Отмены провайдерами', report_data['summary']['provider_cancellations']],
        ['Злоупотребления', report_data['summary']['abuse_cancellations']],
    ]
    
    for row, (label, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Лист по провайдерам
    ws_providers = wb.create_sheet("По провайдерам")
    headers = ['Провайдер', 'Всего отмен', 'Клиентами', 'Провайдерами', 'Злоупотребления']
    for col, header in enumerate(headers, 1):
        cell = ws_providers.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, provider in enumerate(report_data['by_provider'], 2):
        ws_providers.cell(row=row, column=1, value=provider['booking__provider__name'])
        ws_providers.cell(row=row, column=2, value=provider['total_cancellations'])
        ws_providers.cell(row=row, column=3, value=provider['client_cancellations'])
        ws_providers.cell(row=row, column=4, value=provider['provider_cancellations'])
        ws_providers.cell(row=row, column=5, value=provider['abuse_cancellations'])
    
    # Лист детализации
    ws_details = wb.create_sheet("Детализация")
    headers = ['ID бронирования', 'Провайдер', 'Услуга', 'Отменил', 'Причина', 'Злоупотребление', 'Дата']
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row, detail in enumerate(report_data['details'], 2):
        ws_details.cell(row=row, column=1, value=detail['booking_id'])
        ws_details.cell(row=row, column=2, value=detail['provider_name'])
        ws_details.cell(row=row, column=3, value=detail['service_name'])
        ws_details.cell(row=row, column=4, value=detail['cancelled_by'])
        ws_details.cell(row=row, column=5, value=detail['reason'])
        ws_details.cell(row=row, column=6, value='Да' if detail['is_abuse'] else 'Нет')
        ws_details.cell(row=row, column=7, value=detail['created_at'].strftime('%Y-%m-%d %H:%M'))
    
    # Настройка ширины столбцов
    for ws in [ws_summary, ws_providers, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cancellation_report.xlsx'
    wb.save(response)
    return response 