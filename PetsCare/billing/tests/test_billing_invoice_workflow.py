"""
Тесты ручного и автоматического инвойсинга, PDF и FIFO-проводки платежей.
"""

from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.core.management import call_command
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from django.test import RequestFactory
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from billing.invoice_services import InvoiceGenerationService, InvoicePdfService
from billing.models import (
    BillingConfig,
    Currency,
    Invoice,
    InvoiceLine,
    Payment,
    PlatformCompany,
    VATRate,
)
from billing.serializers import InvoiceSerializer
from billing.tasks import generate_scheduled_invoices
from booking.constants import COMPLETED_BY_SYSTEM
from booking.models import Booking, BookingPayment, BookingStatus
from custom_admin import custom_admin_site
from catalog.models import Service
from pets.models import Pet
from production_calendar.models import DAY_TYPE_HOLIDAY, ProductionCalendar
from providers.models import Employee, EmployeeProvider, Provider
from users.models import User


class BillingInvoiceWorkflowTest(TestCase):
    """
    Проверяет ключевые сценарии нового billing workflow.
    """

    @classmethod
    def setUpTestData(cls):
        call_command('generate_billing_data')
        cls.admin_user = User.objects.get(email='billing-admin@example.com')

    def test_generate_invoice_for_online_completed_booking_creates_pdf(self):
        """
        Ручная генерация счета создает invoice, line и PDF-файл.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        booking = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('125.00'),
            completed_days_ago=2,
        )

        service = InvoiceGenerationService()
        invoice = service.generate_for_provider(
            provider=provider,
            start_date=timezone.now().date() - timedelta(days=7),
            end_date=timezone.now().date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None
        self.assertTrue(invoice.lines.filter(booking=booking).exists())
        self.assertIsNotNone(invoice.platform_company)
        self.assertTrue(invoice.pdf_file.name.endswith('.pdf'))
        self.assertGreater(invoice.amount, Decimal('0.00'))
        self.assertEqual(invoice.status, 'sent')

    def test_completed_payment_without_invoice_uses_fifo_for_oldest_open_invoices(self):
        """
        Платеж без выбранного invoice сначала гасит самый старый открытый счет.
        """
        provider = Provider.objects.get(name='Provider_Level2')
        oldest_invoice = provider.invoices.order_by('issued_at', 'created_at').first()
        self.assertIsNotNone(oldest_invoice)
        assert oldest_invoice is not None

        self._create_online_completed_booking(
            provider=provider,
            price=Decimal('140.00'),
            completed_days_ago=1,
        )
        new_invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=timezone.now().date() - timedelta(days=7),
            end_date=timezone.now().date(),
        )

        self.assertIsNotNone(new_invoice)
        assert new_invoice is not None

        oldest_outstanding = oldest_invoice.outstanding_amount
        second_outstanding_before = new_invoice.outstanding_amount

        payment = Payment.objects.create(
            provider=provider,
            amount=oldest_outstanding + Decimal('1.00'),
            status='completed',
            payment_method='bank_transfer',
        )

        oldest_invoice.refresh_from_db()
        new_invoice.refresh_from_db()
        payment.refresh_from_db()

        self.assertIsNotNone(payment.applied_at)
        self.assertEqual(oldest_invoice.status, 'paid')
        self.assertEqual(new_invoice.status, 'partially_paid')
        self.assertEqual(
            new_invoice.outstanding_amount,
            second_outstanding_before - Decimal('1.00'),
        )

    def test_platform_company_is_seeded_by_migration(self):
        """
        После миграций доступно fallback-юрлицо платформы.
        """
        self.assertTrue(PlatformCompany.objects.filter(is_active=True).exists())

    def test_invoice_pdf_renders_period_summary_without_booking_numbers(self):
        """
        PDF-шаблон и fallback показывают сводку за период, а не список booking IDs.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        booking_one = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('125.00'),
            completed_days_ago=11,
        )
        booking_two = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('175.00'),
            completed_days_ago=11,
        )

        invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=booking_one.completed_at.date(),
            end_date=booking_one.completed_at.date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None

        pdf_service = InvoicePdfService()
        context = pdf_service._build_context(invoice, invoice.platform_company)
        html_content = render_to_string(pdf_service.template_name, context)
        fallback_text = '\n'.join(pdf_service._build_fallback_lines(context))

        self.assertIn('Completed online bookings', html_content)
        self.assertIn(str(invoice.start_date), html_content)
        self.assertIn(str(invoice.end_date), html_content)
        self.assertNotIn(f'#{booking_one.pk}', html_content)
        self.assertNotIn(f'#{booking_two.pk}', html_content)
        self.assertIn('Booking count: 2', fallback_text)
        self.assertNotIn('Booking #', fallback_text)

    def test_invoice_admin_booking_breakdown_page_and_excel_export_show_invoice_bookings(self):
        """
        Django admin дает детализацию бронирований по счету и экспорт в Excel.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        booking_one = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('140.00'),
            completed_days_ago=13,
        )
        booking_two = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('160.00'),
            completed_days_ago=13,
        )
        invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=booking_one.completed_at.date(),
            end_date=booking_one.completed_at.date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None

        self.client.force_login(self.admin_user)
        breakdown_url = reverse(
            f'{custom_admin_site.name}:billing_invoice_booking_breakdown',
            args=[invoice.pk],
        )
        response = self.client.get(breakdown_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, booking_one.code)
        self.assertContains(response, booking_two.code)
        self.assertContains(
            response,
            reverse(f'{custom_admin_site.name}:booking_booking_change', args=[booking_one.pk]),
        )
        self.assertContains(response, 'Export to Excel')

        export_response = self.client.get(breakdown_url, {'export': 'xlsx'})

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        first_data_row = [cell.value for cell in worksheet[2]]

        self.assertEqual(headers[:4], ['Booking ID', 'Booking Code', 'Completed At', 'Service'])
        self.assertEqual(first_data_row[0], booking_one.pk)
        self.assertEqual(first_data_row[1], booking_one.code)

    def test_generate_invoice_applies_vat_only_for_non_vat_payer(self):
        """
        При генерации НДС включается только для неплательщика НДС.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        provider.country = 'DE'
        provider.is_vat_payer = False
        provider.save(update_fields=['country', 'is_vat_payer'])
        VATRate.objects.update_or_create(
            country='DE',
            effective_date=timezone.now().date() - timedelta(days=1),
            defaults={'rate': Decimal('19.00'), 'is_active': True},
        )

        booking = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('200.00'),
            completed_days_ago=15,
        )
        invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=booking.completed_at.date(),
            end_date=booking.completed_at.date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None
        line = invoice.lines.get(booking=booking)

        self.assertEqual(line.vat_rate, Decimal('19.00'))
        self.assertGreater(line.vat_amount, Decimal('0.00'))

    def test_provider_admin_api_returns_invoice_booking_breakdown_and_excel_export(self):
        """
        Кабинет провайдера получает детализацию бронирований счета и xlsx-выгрузку.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        provider_admin = self._create_provider_admin(provider)
        booking_one = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('130.00'),
            completed_days_ago=19,
        )
        booking_two = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('170.00'),
            completed_days_ago=19,
        )
        invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=booking_one.completed_at.date(),
            end_date=booking_one.completed_at.date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None

        self.client.force_login(provider_admin)
        response = self.client.get(f'/api/v1/invoices/{invoice.pk}/booking-breakdown/')

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['invoice_id'], invoice.pk)
        self.assertEqual(payload['booking_count'], 2)
        self.assertEqual(payload['rows'][0]['booking_code'], booking_one.code)
        self.assertEqual(payload['rows'][1]['booking_code'], booking_two.code)
        self.assertEqual(payload['currency_code'], invoice.currency.code)

        export_response = self.client.get(
            f'/api/v1/invoices/{invoice.pk}/booking-breakdown/',
            {'export': 'xlsx'},
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(
            export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(export_response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A2'].value, booking_one.pk)
        self.assertEqual(worksheet['B2'].value, booking_one.code)

    def test_generate_scheduled_invoice_for_previous_month_on_configured_working_day(self):
        """
        Автогенерация выставляет счет за предыдущий месяц в нужный рабочий день.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        self._configure_provider_billing_schedule(
            provider,
            invoice_working_day=3,
            payment_working_day=5,
        )
        completed_at = timezone.make_aware(datetime(2026, 3, 20, 12, 0))
        booking = self._create_online_completed_booking_at(
            provider=provider,
            price=Decimal('185.00'),
            completed_at=completed_at,
        )

        invoice = InvoiceGenerationService().generate_scheduled_for_provider(
            provider,
            run_date=date(2026, 4, 3),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None
        invoice.refresh_from_db()

        self.assertEqual(invoice.start_date, date(2026, 3, 1))
        self.assertEqual(invoice.end_date, date(2026, 3, 31))
        self.assertEqual(invoice.issued_at.date(), date(2026, 4, 3))
        self.assertEqual(invoice.payment_record.due_date, date(2026, 4, 7))
        self.assertTrue(invoice.lines.filter(booking=booking).exists())
        self.assertIsNone(
            InvoiceGenerationService().generate_scheduled_for_provider(
                provider,
                run_date=date(2026, 4, 3),
            )
        )

    def test_scheduled_invoice_task_uses_production_calendar_holiday_override(self):
        """
        Daily task сдвигает дату счета на следующий рабочий день по ProductionCalendar.
        """
        provider = Provider.objects.get(name='Provider_Level1')
        self._configure_provider_billing_schedule(
            provider,
            invoice_working_day=1,
            payment_working_day=3,
        )
        ProductionCalendar.objects.update_or_create(
            country='ME',
            date=date(2026, 6, 1),
            defaults={
                'day_type': DAY_TYPE_HOLIDAY,
                'description': 'Synthetic billing holiday',
                'is_manually_corrected': True,
            },
        )
        booking = self._create_online_completed_booking_at(
            provider=provider,
            price=Decimal('210.00'),
            completed_at=timezone.make_aware(datetime(2026, 5, 18, 14, 0)),
        )

        self.assertFalse(provider.should_generate_invoice_on(date(2026, 6, 1)))
        self.assertTrue(provider.should_generate_invoice_on(date(2026, 6, 2)))

        skipped_result = generate_scheduled_invoices(run_date_iso='2026-06-01')
        self.assertEqual(skipped_result['status'], 'completed')
        self.assertEqual(skipped_result['statistics']['generated_invoices'], 0)

        task_result = generate_scheduled_invoices(run_date_iso='2026-06-02')
        self.assertEqual(task_result['status'], 'completed')
        self.assertEqual(task_result['statistics']['generated_invoices'], 1)

        invoice = Invoice.objects.get(
            provider=provider,
            start_date=date(2026, 5, 1),
            end_date=date(2026, 5, 31),
        )
        self.assertEqual(invoice.issued_at.date(), date(2026, 6, 2))
        self.assertEqual(invoice.payment_record.due_date, date(2026, 6, 4))
        self.assertTrue(invoice.lines.filter(booking=booking).exists())

    def test_scheduled_invoice_task_catches_up_after_missed_issue_day(self):
        """
        Если daily task пропустила расчетный рабочий день, следующий запуск
        все равно создаст счет за прошлый месяц.
        """
        provider = Provider.objects.get(name='Provider_Level2')
        self._reset_provider_invoices(provider)
        self._configure_provider_billing_schedule(
            provider,
            invoice_working_day=3,
            payment_working_day=8,
        )
        booking = self._create_online_completed_booking_at(
            provider=provider,
            price=Decimal('199.00'),
            completed_at=timezone.make_aware(datetime(2026, 3, 10, 11, 0)),
        )

        self.assertFalse(provider.should_generate_invoice_on(date(2026, 4, 6)))

        task_result = generate_scheduled_invoices(run_date_iso='2026-04-06')
        self.assertEqual(task_result['status'], 'completed')
        self.assertEqual(task_result['statistics']['generated_invoices'], 1)

        invoice = Invoice.objects.get(
            provider=provider,
            start_date=date(2026, 3, 1),
            end_date=date(2026, 3, 31),
        )
        self.assertEqual(invoice.issued_at.date(), date(2026, 4, 6))
        self.assertEqual(invoice.payment_record.due_date, provider.calculate_invoice_due_date(date(2026, 4, 6)))
        self.assertTrue(invoice.lines.filter(booking=booking).exists())

    def test_scheduled_invoice_task_catches_up_multiple_missed_months(self):
        """
        Один запуск задачи может догенерировать несколько просроченных месяцев.
        """
        provider = Provider.objects.get(name='Provider_Level3')
        self._reset_provider_invoices(provider)
        self._configure_provider_billing_schedule(
            provider,
            invoice_working_day=2,
            payment_working_day=6,
        )
        february_booking = self._create_online_completed_booking_at(
            provider=provider,
            price=Decimal('101.00'),
            completed_at=timezone.make_aware(datetime(2026, 2, 14, 10, 0)),
        )
        march_booking = self._create_online_completed_booking_at(
            provider=provider,
            price=Decimal('151.00'),
            completed_at=timezone.make_aware(datetime(2026, 3, 19, 15, 0)),
        )

        task_result = generate_scheduled_invoices(run_date_iso='2026-04-10')
        self.assertEqual(task_result['status'], 'completed')
        self.assertEqual(task_result['statistics']['generated_invoices'], 2)

        february_invoice = Invoice.objects.get(
            provider=provider,
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
        )
        march_invoice = Invoice.objects.get(
            provider=provider,
            start_date=date(2026, 3, 1),
            end_date=date(2026, 3, 31),
        )

        self.assertEqual(february_invoice.issued_at.date(), date(2026, 4, 10))
        self.assertEqual(march_invoice.issued_at.date(), date(2026, 4, 10))
        self.assertTrue(february_invoice.lines.filter(booking=february_booking).exists())
        self.assertTrue(march_invoice.lines.filter(booking=march_booking).exists())

    def test_download_pdf_endpoint_regenerates_existing_invoice_pdf(self):
        """
        Скачивание PDF через API не отдает устаревший сохраненный файл.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        provider_admin = self._create_provider_admin(provider)
        booking = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('145.00'),
            completed_days_ago=21,
        )
        invoice = InvoiceGenerationService().generate_for_provider(
            provider=provider,
            start_date=booking.completed_at.date(),
            end_date=booking.completed_at.date(),
        )

        self.assertIsNotNone(invoice)
        assert invoice is not None

        invoice.pdf_file.save('stale-invoice.pdf', ContentFile(b'OLD PDF CONTENT'), save=True)
        self.client.force_login(provider_admin)

        with patch(
            'billing.invoice_services.InvoicePdfService._render_pdf_bytes',
            return_value=b'REGENERATED PDF CONTENT',
        ):
            response = self.client.get(f'/api/v1/invoices/{invoice.pk}/download-pdf/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        response_bytes = b''.join(response.streaming_content)
        self.assertEqual(response_bytes, b'REGENERATED PDF CONTENT')

    def test_invoice_line_save_clears_vat_for_vat_payer(self):
        """
        Даже при ручном сохранении строки счета НДС не остается у VAT payer.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        provider.country = 'DE'
        provider.is_vat_payer = True
        provider.save(update_fields=['country', 'is_vat_payer'])
        currency = provider.invoice_currency or Currency.objects.get(code='EUR')
        booking = self._create_online_completed_booking(
            provider=provider,
            price=Decimal('90.00'),
            completed_days_ago=17,
        )
        invoice = Invoice.objects.create(
            provider=provider,
            platform_company=PlatformCompany.resolve_for_provider(provider),
            start_date=booking.completed_at.date(),
            end_date=booking.completed_at.date(),
            amount=Decimal('0.00'),
            currency=currency,
            status='sent',
            issued_at=timezone.now(),
        )

        line = InvoiceLine.objects.create(
            invoice=invoice,
            booking=booking,
            amount=Decimal('90.00'),
            commission=Decimal('9.00'),
            rate=Decimal('10.00'),
            currency=currency,
            vat_rate=Decimal('19.00'),
            vat_amount=Decimal('1.71'),
            total_with_vat=Decimal('10.71'),
        )
        line.refresh_from_db()

        self.assertIsNone(line.vat_rate)
        self.assertEqual(line.vat_amount, Decimal('0.00'))
        self.assertEqual(line.total_with_vat, Decimal('9.00'))

    def test_invoice_serializer_builds_pdf_download_url_without_namespace_reverse(self):
        """
        Сериализация счета не падает из-за namespace reverse и отдает абсолютный URL PDF.
        """
        provider = Provider.objects.get(name='Provider_FullyPaid')
        invoice = provider.invoices.order_by('id').first()

        self.assertIsNotNone(invoice)
        assert invoice is not None

        request = RequestFactory().get('/api/v1/invoices/')
        request.META['HTTP_HOST'] = 'testserver'
        serializer = InvoiceSerializer(invoice, context={'request': request})

        self.assertEqual(
            serializer.data['pdf_download_url'],
            f'http://testserver/api/v1/invoices/{invoice.id}/download-pdf/',
        )

    def _create_online_completed_booking(self, provider, price, completed_days_ago):
        """
        Создает completed booking с online payment для тестов инвойсинга.
        """
        completed_at = timezone.now() - timedelta(days=completed_days_ago)
        return self._create_online_completed_booking_at(
            provider=provider,
            price=price,
            completed_at=completed_at,
        )

    def _create_online_completed_booking_at(self, provider, price, completed_at):
        """
        Создает completed booking с online payment в конкретный момент времени.
        """
        owner = User.objects.get(email='billing-demo-owner@example.com')
        pet = Pet.objects.get(name='Billing Demo Pet')
        employee = Employee.objects.filter(providers=provider).first()
        service = Service.objects.get(code='billing_demo_service')
        location = provider.locations.first()
        status, _ = BookingStatus.objects.get_or_create(name='completed')
        if timezone.is_naive(completed_at):
            completed_at = timezone.make_aware(completed_at)

        booking = Booking.objects.create(
            user=owner,
            employee=employee,
            provider_location=location,
            service=service,
            pet=pet,
            start_time=completed_at - timedelta(hours=1),
            end_time=completed_at,
            status=status,
            price=price,
            completed_at=completed_at,
            completed_by_actor=COMPLETED_BY_SYSTEM,
        )
        BookingPayment.objects.create(
            booking=booking,
            amount=price,
            payment_method='online',
            transaction_id=f'ONLINE-{booking.pk}',
        )
        return booking

    def _configure_provider_billing_schedule(self, provider, *, invoice_working_day, payment_working_day):
        """
        Привязывает провайдера к billing config с нужным рабочим днем счета и оплаты.
        """
        currency = provider.invoice_currency or Currency.objects.get(code='EUR')
        provider.country = 'ME'
        provider.invoice_currency = currency
        provider.save(update_fields=['country', 'invoice_currency'])

        acceptance = provider.document_acceptances.select_related('document__billing_config').get(
            document__document_type__code='global_offer',
            is_active=True,
        )
        document = acceptance.document
        billing_config = document.billing_config
        if billing_config is None:
            billing_config = BillingConfig.objects.create(
                name=f'Test Billing Config {provider.pk}',
                commission_percent=Decimal('10.00'),
                payment_deferral_days=payment_working_day,
                invoice_period_days=invoice_working_day,
                is_active=True,
            )
            document.billing_config = billing_config
            document.save(update_fields=['billing_config'])
            return

        billing_config.payment_deferral_days = payment_working_day
        billing_config.invoice_period_days = invoice_working_day
        billing_config.save(update_fields=['payment_deferral_days', 'invoice_period_days', 'updated_at'])

    def _create_provider_admin(self, provider):
        """
        Создает provider admin, привязанного к нужному провайдеру.
        """
        provider_admin = User.objects.create_user(
            email=f'provider-admin-{provider.pk}@example.com',
            password='Secret123!',
            username=f'provider-admin-{provider.pk}',
            phone_number=f'+1202555{provider.pk:04d}',
        )
        provider_admin.add_role('provider_admin')
        employee = Employee.objects.create(user=provider_admin, is_active=True)
        EmployeeProvider.objects.create(
            employee=employee,
            provider=provider,
            role=EmployeeProvider.ROLE_PROVIDER_ADMIN,
            is_provider_admin=True,
            start_date=timezone.now().date(),
        )
        return provider_admin

    def _reset_provider_invoices(self, provider):
        """
        Очищает счета и блокировки провайдера, чтобы тесты scheduled billing
        проверяли только созданные ими периоды.
        """
        provider.blockings.all().delete()
        Payment.objects.filter(provider=provider).delete()
        Invoice.objects.filter(provider=provider).delete()
